from flask import Flask, send_from_directory, send_file, render_template, request, jsonify, session, abort, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os, re, sys, io, tempfile

# Make BlueprintLinker importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'BlueprintLinker'))

app = Flask(__name__)
app.secret_key = "pei-tools-secret-2024"
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Behind the nginx proxy, trust X-Forwarded-Proto/Host so generated absolute
# links (public links, QR codes, task-complete links) say https://peitools.com
# instead of http:// — without this, every QR scan takes a redirect hop.
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)


# ── Jobs folder (persisted via Docker volume) ────────────────
JOBS_DIR = os.path.join(os.path.dirname(__file__), "jobs")
os.makedirs(JOBS_DIR, exist_ok=True)

ADMIN_PASSWORD = "PEI2024"

# ── User accounts ─────────────────────────────────────────────
_USERS_RAW = {
    "dennisa@pacificerectors.com":  "Andersen4460",
    "armandor@pacificerectors.com": "Rivera4460",
    "arturov@pacificerectors.com":  "Vargas4460",
    "caseyc@pacificerectors.com":   "Carlson4460",
    "davida@pacificerectors.com":   "Arias4460",
    "elim@pacificerectors.com":     "Martinez4460",
    "erica@pacificerectors.com":    "Andersen4460",
    "glenw@pacificerectors.com":    "Wheeler4460",
    "gustavoh@pacificerectors.com": "Hernandez4460",
    "javierm@pacificerectors.com":  "Arevalo4460",
    "juanc@pacificerectors.com":    "Camarena4460",
    "luism@pacificerectors.com":    "Marure4460",
    "robinp@pacificerectors.com":   "Pederson4460",
    "stevens@pacificerectors.com":  "Sousa4460",
    "thomasr@pacificerectors.com":  "Rowley4460",
    "tommyp@pacificerectors.com":   "Pearman4460",
    "miket@pacificerectors.com":    "Thomas4460",
    "jeffy@pacificerectors.com":    "Young4460",
    "jasonw@pacificerectors.com":   "Walters4460",
    "dalynb@pacificerectors.com":   "Bush4460",
    "thomasm@pacificerectors.com":  "McClelland4460",
    "fritzb@pacificerectors.com":   "Bowen4460",
    "erics@pacificerectors.com":    "Sidener4460",
    "debid@pacificerectors.com":    "Dunkin4460",
    "kellyl@pacificerectors.com":   "Lee4460",
    "kishag@pacificerectors.com":   "Gann4460",
    "jennab@pacificerectors.com":   "Bearden4460",
    "meganf@pacificerectors.com":   "Friery4460",
    "nicholaskoron@gmail.com":      "REDFredf",
}
USERS = {email: generate_password_hash(pwd) for email, pwd in _USERS_RAW.items()}

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user"):
            # API routes get a JSON 401; page routes get a redirect to login
            is_json = (request.content_type or '').startswith('application/json')
            if request.path.startswith("/api/") or request.path.startswith("/files/") or is_json:
                return jsonify({"error": "Unauthorized — please log in again"}), 401
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return decorated

# ── Login / Logout ─────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        hashed = USERS.get(email)
        if hashed and check_password_hash(hashed, password):
            session["user"] = email
            return redirect(request.args.get("next") or url_for("index"))
        error = "Incorrect email or password."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))

CATEGORIES = ["Blueprints", "Packing Lists", "Fab Sheets", "Panel Mapper", "Spreadsheets"]
CAD_FOLDER = "DXF CAD FILE"
ALL_FOLDERS = CATEGORIES + [CAD_FOLDER]
IMAGE_EXTS = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff", ".tif"}

def safe_join(*parts):
    """Resolve a path and ensure it stays inside JOBS_DIR."""
    path = os.path.realpath(os.path.join(JOBS_DIR, *parts))
    if not path.startswith(os.path.realpath(JOBS_DIR)):
        abort(403)
    return path

# ── Landing page ─────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    return render_template("index.html")

# ── Existing tools ───────────────────────────────────────────
@app.route("/sheet_editor")
@login_required
def sheet_editor():
    return render_template("sheet_editor.html")

@app.route("/sheet_extractor", methods=["GET", "POST"])
@login_required
def sheet_extractor():
    if request.method == "GET":
        return render_template("panel_sheet_mapper.html")

    files = request.files.getlist("files")
    if not files:
        return "No files uploaded", 400

    mapping = {}

    for f in files:
        fname = f.filename
        lines = [l.decode("utf-8", errors="ignore").strip() for l in f.read().splitlines()]

        layout_handles = []
        for i in range(1, len(lines)-1):
            if lines[i] == "AcDbLayout" and lines[i-1] == "100":
                name = handle = None
                j = i + 1
                while j < min(i+160, len(lines)-1):
                    c, v = lines[j], lines[j+1] if j+1 < len(lines) else ""
                    if c == "1" and name is None: name = v
                    if c == "330": handle = v
                    if c == "0" and j > i+2: break
                    j += 2
                if name and name != "Model":
                    layout_handles.append((name, handle))

        owner_to_block = {}
        for i in range(0, len(lines)-3, 2):
            if lines[i] == "0" and lines[i+1] == "BLOCK":
                bname = owner = None
                for j in range(i+2, min(i+30, len(lines)-1), 2):
                    if lines[j] == "2": bname = lines[j+1]
                    if lines[j] == "330": owner = lines[j+1]
                    if lines[j] == "0" and j > i+2: break
                if bname and owner:
                    owner_to_block[owner] = bname

        layout_to_block = {name: owner_to_block.get(handle, "") for name, handle in layout_handles}
        block_to_layout = {v: k for k, v in layout_to_block.items() if v}

        current_block = None
        for i in range(0, len(lines)-3, 2):
            if lines[i] == "0" and lines[i+1] == "BLOCK":
                for j in range(i+2, min(i+20, len(lines)-1), 2):
                    if lines[j] == "2": current_block = lines[j+1]; break
            elif lines[i] == "0" and lines[i+1] == "ENDBLK":
                current_block = None
            elif lines[i] == "0" and lines[i+1] in ("TEXT", "MTEXT"):
                layer = text = None
                for j in range(i+2, min(i+40, len(lines)-1), 2):
                    c, v = lines[j], lines[j+1] if j+1 < len(lines) else ""
                    if c == "8": layer = v
                    if c == "1": text = v
                    if c == "0" and j > i+2: break
                if layer and "PANEL" in layer.upper() and text and text.strip().isdigit():
                    n = int(text.strip())
                    sheet = block_to_layout.get(current_block)
                    if sheet:
                        if n not in mapping:
                            mapping[n] = {"sheets": [], "file": fname}
                        if sheet not in mapping[n]["sheets"]:
                            mapping[n]["sheets"].append(sheet)

    def sort_sheets(sheets):
        def key(s):
            try: return [float(p) for p in s.split(".")]
            except: return [999]
        return sorted(sheets, key=key)

    for p in mapping:
        mapping[p]["sheets"] = sort_sheets(mapping[p]["sheets"])

    str_mapping = {str(k): v for k, v in sorted(mapping.items())}
    return jsonify({"mapping": str_mapping, "job_name": "", "count": len(str_mapping)})

# ── CAD file listing ─────────────────────────────────────────
@app.route("/api/jobs/<path:job>/dxf-files")
def api_dxf_files(job):
    cad_path = safe_join(job, CAD_FOLDER)
    if not os.path.isdir(cad_path):
        return jsonify([])
    files = sorted([f for f in os.listdir(cad_path) if f.lower().endswith(".dxf")])
    return jsonify(files)

@app.route("/cad-files/<path:filepath>")
@login_required
def serve_cad_file(filepath):
    full = safe_join(filepath)
    return send_from_directory(os.path.dirname(full), os.path.basename(full))

# ── Document Viewer (public) ─────────────────────────────────
@app.route("/blueprints")
@app.route("/documents")
@login_required
def blueprints():
    return render_template("blueprints.html")

@app.route("/field-compass")
@login_required
def field_compass():
    return render_template("field_compass.html")

@app.route("/api/jobs")
def api_jobs():
    if not os.path.isdir(JOBS_DIR):
        return jsonify([])
    jobs = sorted([d for d in os.listdir(JOBS_DIR)
                   if os.path.isdir(os.path.join(JOBS_DIR, d))])
    return jsonify(jobs)

@app.route("/api/jobs/<path:job>")
def api_job_files(job):
    import time
    job_path = safe_join(job)
    if not os.path.isdir(job_path):
        return jsonify({"error": "Job not found"}), 404

    def file_entry(folder, fname):
        fpath = os.path.join(folder, fname)
        try:
            mtime = os.path.getmtime(fpath)
            modified = time.strftime("%-m/%-d/%y", time.localtime(mtime))
        except Exception:
            modified = ""
        return {"name": fname, "modified": modified}

    result = {}
    for cat in CATEGORIES:
        cat_path = os.path.join(job_path, cat)
        if os.path.isdir(cat_path):
            if cat == "Spreadsheets":
                exts = {".xlsx", ".xls"}
            else:
                exts = {".pdf"}
            files = sorted([f for f in os.listdir(cat_path)
                            if os.path.splitext(f)[1].lower() in exts])
            if files:
                result[cat] = [file_entry(cat_path, f) for f in files]
    # Include DXF CAD FILE so admin can verify uploads
    cad_path = os.path.join(job_path, CAD_FOLDER)
    if os.path.isdir(cad_path):
        cad_files = sorted([f for f in os.listdir(cad_path)
                            if os.path.splitext(f)[1].lower() in {".dxf", ".dwg"}])
        if cad_files:
            result[CAD_FOLDER] = [file_entry(cad_path, f) for f in cad_files]
    return jsonify(result)

@app.route("/api/jobs/<path:job>/all-files")
def api_job_all_files(job):
    """Return all viewable files (PDF + images) across all categories."""
    job_path = safe_join(job)
    if not os.path.isdir(job_path):
        return jsonify({"error": "Job not found"}), 404
    files = []
    for cat in CATEGORIES:
        cat_path = os.path.join(job_path, cat)
        if os.path.isdir(cat_path):
            for f in sorted(os.listdir(cat_path)):
                ext = os.path.splitext(f)[1].lower()
                if ext in IMAGE_EXTS:
                    files.append({"name": f, "category": cat,
                                  "url": f"/files/{job}/{cat}/{f}",
                                  "type": "pdf" if ext == ".pdf" else "image"})
    return jsonify(files)

def _extract_page_drawing_numbers(pdf_path, page_indices):
    """
    Extract KPS drawing numbers (dep. no. field, e.g. "2.3") from each page.
    Strategy: pull all words from the right 20% of the page (KPS title block column),
    find every word that is purely in X.Y / X.YY format, then pick the one
    physically closest to the bottom-right corner — that is always dep. no.
    Falls back to str(page_index + 1) if nothing is found.
    """
    import re as _re_dn
    result = {}
    try:
        import fitz
        doc = fitz.open(pdf_path)
        for page_num in sorted(page_indices):
            if page_num < 0 or page_num >= doc.page_count:
                result[page_num] = str(page_num + 1)
                continue
            page = doc[page_num]
            w, h = page.rect.width, page.rect.height

            # Right 20% of page = KPS title block column
            clip = fitz.Rect(w * 0.80, 0, w, h)
            # words: (x0, y0, x1, y1, text, block_no, line_no, word_no)
            words = page.get_text("words", clip=clip)

            candidates = []
            for word in words:
                text = word[4].strip()
                # Must be exactly X.Y or X.YY — no extra chars (filters "2/3.4", dates, etc.)
                if _re_dn.match(r'^\d{1,3}\.\d{1,2}$', text):
                    cx = (word[0] + word[2]) / 2
                    cy = (word[1] + word[3]) / 2
                    dist = ((cx - w) ** 2 + (cy - h) ** 2) ** 0.5
                    candidates.append((dist, text))

            if candidates:
                candidates.sort()                    # closest to bottom-right first
                result[page_num] = candidates[0][1]
            else:
                result[page_num] = str(page_num + 1)

        doc.close()
    except Exception:
        for p in page_indices:
            result.setdefault(p, str(p + 1))
    return result


@app.route("/api/jobs/<path:job_name>/build-spreadsheet", methods=["POST"])
@login_required
def build_spreadsheet(job_name):
    import json as _json_ss, time as _time_ss, re as _re_ss
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl not installed — run deploy.bat to rebuild Docker image"}), 500

    job_path = safe_join(job_name)
    if not os.path.isdir(job_path):
        return jsonify({"error": "Job not found"}), 404

    # ── Load delivery state ───────────────────────────────────
    state_path = _pl_state_path(job_name)
    if not os.path.exists(state_path):
        return jsonify({"error": "No packing list data found. Process a packing list in the Packing List Tracker first."}), 400

    with open(state_path) as f:
        delivery_state = _json_ss.load(f)

    # ── Load panel locations (Panel Map preferred, fallback to PL cache) ──
    panel_locs = {}
    pm_sess = None
    try:
        pm_sess = _pm_load_session(job_name)
        if pm_sess:
            locs_path = pm_sess.get("locs", "")
            if locs_path and os.path.isfile(locs_path):
                with open(locs_path) as f:
                    panel_locs = _json_ss.load(f)
    except Exception:
        pass
    if not panel_locs:
        cache_path = _pl_cache_path(job_name)
        if os.path.exists(cache_path):
            with open(cache_path) as f:
                panel_locs = _json_ss.load(f)

    # ── Extract drawing numbers from title block ──────────────────
    # scan_pdf is the trimmed PDF whose page indices match panel_locs exactly.
    # src_pdf is the full blueprint — page indices do NOT match panel_locs.
    drawing_nums = {}
    if panel_locs and pm_sess:
        try:
            src_pdf = pm_sess.get("scan_pdf") or pm_sess.get("src_pdf", "")
            if src_pdf and os.path.isfile(src_pdf):
                unique_pages = {loc["page"] for loc in panel_locs.values()
                                if loc.get("page") is not None}
                drawing_nums = _extract_page_drawing_numbers(src_pdf, unique_pages)
        except Exception:
            pass

    # ── Get packing list file dates for "Date Delivered" ──────
    pl_dir = safe_join(job_name, "Packing Lists")
    shipment_dates = {}
    if os.path.isdir(pl_dir):
        for fname in os.listdir(pl_dir):
            if fname.lower().endswith(".pdf"):
                label = fname[:-4]  # strip .pdf
                fpath = os.path.join(pl_dir, fname)
                try:
                    mtime = os.path.getmtime(fpath)
                    shipment_dates[label] = _time_ss.strftime("%-m/%-d/%Y", _time_ss.localtime(mtime))
                except Exception:
                    shipment_dates[label] = ""

    # ── Build spreadsheet ──────────────────────────────────────
    ss_dir = safe_join(job_name, "Spreadsheets")
    os.makedirs(ss_dir, exist_ok=True)
    out_path = os.path.join(ss_dir, f"{job_name}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Panel Data"

    headers = ["Panel Number", "Sheet Number", "Order Number", "Packing List", "Date Delivered"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    def _panel_sort_key(display_name):
        m = _re_ss.match(r"(\d+)", str(display_name))
        return (int(m.group(1)) if m else float('inf'), str(display_name))

    # Build rows from Panel Mapper locs (source of panels_only PDF); fallback to delivery_state
    if panel_locs:
        # Each entry in panel_locs: key may be "211" or "211#2" (duplicate); loc has page, optional label
        rows = []
        for key, loc in panel_locs.items():
            display_name = loc.get("label") or _re_ss.sub(r'#\d+$', '', key)
            page_num = loc.get("page")
            sheet_num = drawing_nums.get(page_num, page_num + 1) if page_num is not None else ""
            info = delivery_state.get(display_name) or delivery_state.get(str(display_name)) or {}
            shipment = info.get("shipment", "")
            order_num = info.get("order_num", "")
            date_del = shipment_dates.get(shipment, "")
            rows.append((display_name, sheet_num, order_num, shipment, date_del))
        rows.sort(key=lambda r: _panel_sort_key(r[0]))
    else:
        # Fallback: use delivery_state keys
        rows = []
        for panel in sorted(delivery_state.keys(), key=_panel_sort_key):
            info = delivery_state[panel]
            shipment = info.get("shipment", "")
            order_num = info.get("order_num", "")
            date_del = shipment_dates.get(shipment, "")
            rows.append((panel, "", order_num, shipment, date_del))

    # Apply any persisted manual overrides
    overrides_path = os.path.join(ss_dir, f"{job_name}_overrides.json")
    overrides = {}
    if os.path.exists(overrides_path):
        try:
            with open(overrides_path) as _of:
                overrides = _json_ss.load(_of)
        except Exception:
            overrides = {}
    col_name_to_idx = {h: i for i, h in enumerate(headers)}
    rows_final = []
    for row in rows:
        panel = row[0]
        if panel in overrides:
            row = list(row)
            for col_header, value in overrides[panel].items():
                idx = col_name_to_idx.get(col_header)
                if idx is not None:
                    row[idx] = value
            row = tuple(row)
        rows_final.append(row)

    for row_idx, (panel, sheet_num, order_num, shipment, date_del) in enumerate(rows_final, 2):
        ws.cell(row=row_idx, column=1, value=panel)
        ws.cell(row=row_idx, column=2, value=sheet_num)
        ws.cell(row=row_idx, column=3, value=order_num)
        ws.cell(row=row_idx, column=4, value=shipment)
        ws.cell(row=row_idx, column=5, value=date_del)

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(out_path)
    return jsonify({"ok": True, "panels": len(rows_final), "file": f"{job_name}.xlsx"})


@app.route("/api/jobs/<path:job_name>/spreadsheet-edits", methods=["POST"])
@login_required
def spreadsheet_edits(job_name):
    """Save manual cell edits to overrides JSON and apply to xlsx."""
    import json as _json_ed
    try:
        import openpyxl as _xl_ed
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500
    payload = request.get_json(silent=True) or {}
    edits = payload.get("edits", [])
    if not edits:
        return jsonify({"ok": True, "cells_updated": 0})
    ss_dir = safe_join(job_name, "Spreadsheets")
    out_path = os.path.join(ss_dir, f"{job_name}.xlsx")
    overrides_path = os.path.join(ss_dir, f"{job_name}_overrides.json")
    if not os.path.exists(out_path):
        return jsonify({"error": "Spreadsheet not found — run Pull Data first"}), 404

    # Load existing overrides
    overrides = {}
    if os.path.exists(overrides_path):
        try:
            with open(overrides_path) as f:
                overrides = _json_ed.load(f)
        except Exception:
            overrides = {}

    # Merge new edits into overrides
    for edit in edits:
        panel = str(edit.get("panel", "")).strip()
        col_header = str(edit.get("col_header", "")).strip()
        value = str(edit.get("value", ""))
        if panel and col_header:
            overrides.setdefault(panel, {})[col_header] = value

    # Save overrides JSON
    os.makedirs(ss_dir, exist_ok=True)
    with open(overrides_path, "w") as f:
        _json_ed.dump(overrides, f, indent=2)

    # Also apply to xlsx so it's immediately up-to-date
    wb = _xl_ed.load_workbook(out_path)
    ws = wb.active
    # Build header->col map from row 1
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {str(h): i + 1 for i, h in enumerate(header_row) if h is not None}
    # Build panel->row map from column 1
    panel_row_map = {}
    for r in range(2, ws.max_row + 1):
        pval = ws.cell(row=r, column=1).value
        if pval is not None:
            panel_row_map[str(pval)] = r
    updated = 0
    for edit in edits:
        panel = str(edit.get("panel", "")).strip()
        col_header = str(edit.get("col_header", "")).strip()
        value = str(edit.get("value", ""))
        r = panel_row_map.get(panel)
        c = col_map.get(col_header)
        if r and c:
            ws.cell(row=r, column=c, value=value)
            updated += 1
    wb.save(out_path)
    return jsonify({"ok": True, "cells_updated": updated})


@app.route("/api/jobs/<path:job_name>/sheets-url", methods=["GET"])
@login_required
def get_sheets_url(job_name):
    import json as _json_su
    ss_dir = safe_join(job_name, "Spreadsheets")
    url_path = os.path.join(ss_dir, f"{job_name}_sheets_url.json")
    if os.path.exists(url_path):
        try:
            with open(url_path) as f:
                return jsonify(_json_su.load(f))
        except Exception:
            pass
    return jsonify({"url": None})


@app.route("/api/jobs/<path:job_name>/sheets-url", methods=["POST"])
@login_required
def save_sheets_url(job_name):
    import json as _json_su
    data = request.get_json(silent=True) or {}
    url = data.get("url", "").strip()
    ss_dir = safe_join(job_name, "Spreadsheets")
    os.makedirs(ss_dir, exist_ok=True)
    url_path = os.path.join(ss_dir, f"{job_name}_sheets_url.json")
    with open(url_path, "w") as f:
        _json_su.dump({"url": url or None}, f)
    return jsonify({"ok": True, "url": url or None})


@app.route("/admin", methods=["GET"])
@login_required
def admin():
    return render_template("admin.html")



@app.route("/admin/login", methods=["POST"])
def admin_login():
    data = request.get_json()
    if data and data.get("password") == ADMIN_PASSWORD:
        session["admin"] = True
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Wrong password"}), 401

@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin", None)
    return jsonify({"ok": True})

@app.route("/admin/create-job", methods=["POST"])
@login_required
def admin_create_job():
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    job_name = (data or {}).get("name", "").strip()
    if not job_name or re.search(r'[\\/:*?"<>|]', job_name):
        return jsonify({"error": "Invalid job name"}), 400
    job_path = safe_join(job_name)
    for cat in ALL_FOLDERS:
        os.makedirs(os.path.join(job_path, cat), exist_ok=True)
    return jsonify({"ok": True})

@app.route("/admin/upload", methods=["POST"])
@login_required
def admin_upload():
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 401
    job      = request.form.get("job", "").strip()
    category = request.form.get("category", "").strip()
    if not job or category not in ALL_FOLDERS:
        return jsonify({"error": "Invalid job or category"}), 400
    cat_path = safe_join(job, category)
    os.makedirs(cat_path, exist_ok=True)
    ALLOWED_EXTS = IMAGE_EXTS | {".dxf", ".dwg"}
    uploaded = []
    for f in request.files.getlist("files"):
        ext = os.path.splitext(f.filename)[1].lower()
        if ext in ALLOWED_EXTS:
            fname = re.sub(r'[\\/:*?"<>|]', "_", f.filename)
            f.save(os.path.join(cat_path, fname))
            uploaded.append(fname)
    return jsonify({"ok": True, "uploaded": uploaded})

@app.route("/admin/delete-file", methods=["POST"])
@login_required
def admin_delete_file():
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 401
    data     = request.get_json()
    job      = (data or {}).get("job", "").strip()
    category = (data or {}).get("category", "").strip()
    filename = (data or {}).get("filename", "").strip()
    if not job or category not in CATEGORIES or not filename:
        return jsonify({"error": "Invalid parameters"}), 400
    filepath = safe_join(job, category, filename)
    if os.path.isfile(filepath):
        os.remove(filepath)
    return jsonify({"ok": True})

@app.route("/admin/delete-job", methods=["POST"])
@login_required
def admin_delete_job():
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    job  = (data or {}).get("job", "").strip()
    if not job:
        return jsonify({"error": "Invalid job"}), 400
    import shutil
    job_path = safe_join(job)
    if os.path.isdir(job_path):
        shutil.rmtree(job_path)
    return jsonify({"ok": True})

@app.route("/admin/check-auth")
@login_required
def admin_check_auth():
    return jsonify({"authenticated": bool(session.get("admin"))})

# ── Static files ─────────────────────────────────────────────
@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory(os.path.join(os.path.dirname(__file__), "static"), filename)

@app.route("/files/<path:filepath>")
@login_required
def serve_file(filepath):
    full = safe_join(filepath)
    directory = os.path.dirname(full)
    filename  = os.path.basename(full)
    return send_from_directory(directory, filename)

# ── Share links (job-level) ───────────────────────────────────
import json, secrets
from datetime import datetime, timezone

SHARES_FILE = os.path.join(JOBS_DIR, ".shares.json")

def load_shares():
    if os.path.exists(SHARES_FILE):
        try:
            with open(SHARES_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_shares(shares):
    with open(SHARES_FILE, "w") as f:
        json.dump(shares, f)

@app.route("/api/share", methods=["POST"])
@login_required
def create_share():
    data = request.get_json()
    job = (data or {}).get("job", "").strip()
    if not job:
        return jsonify({"error": "Missing job"}), 400
    shares = load_shares()
    # Remove existing token for this job
    shares = {t: v for t, v in shares.items() if v.get("job") != job}
    token = secrets.token_urlsafe(16)
    shares[token] = {"job": job, "created": datetime.utcnow().isoformat()}
    save_shares(shares)
    return jsonify({"token": token, "url": f"/share/{token}"})

@app.route("/api/share/info", methods=["POST"])
@login_required
def share_info():
    data = request.get_json()
    job = (data or {}).get("job", "").strip()
    shares = load_shares()
    for token, v in shares.items():
        if v.get("job") == job:
            return jsonify({"token": token, "url": f"/share/{token}"})
    return jsonify({"token": None})

@app.route("/api/share/delete", methods=["POST"])
@login_required
def delete_share():
    data = request.get_json()
    token = (data or {}).get("token", "").strip()
    shares = load_shares()
    shares.pop(token, None)
    save_shares(shares)
    return jsonify({"ok": True})

@app.route("/share/<token>")
def view_share(token):
    shares = load_shares()
    info = shares.get(token)
    if not info:
        return "This link has expired or does not exist.", 404
    job = info["job"]
    job_path = os.path.join(JOBS_DIR, job)
    files = {}
    for cat in CATEGORIES:
        cat_path = os.path.join(job_path, cat)
        if os.path.isdir(cat_path):
            cat_files = sorted([f for f in os.listdir(cat_path) if f.lower().endswith(".pdf")])
            if cat_files:
                files[cat] = cat_files
    return render_template("share_view.html", token=token, job=job, files=files)

@app.route("/share/<token>/file/<path:filepath>")
def serve_share_file(token, filepath):
    shares = load_shares()
    info = shares.get(token)
    if not info:
        return "Link not found", 404
    full = safe_join(info["job"], filepath)
    return send_from_directory(os.path.dirname(full), os.path.basename(full))

# ── Compass Share Links ───────────────────────────────────────
COMPASS_SHARES_FILE = os.path.join(JOBS_DIR, ".compass_shares.json")

def load_compass_shares():
    if os.path.exists(COMPASS_SHARES_FILE):
        try:
            with open(COMPASS_SHARES_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_compass_shares(shares):
    with open(COMPASS_SHARES_FILE, "w") as f:
        json.dump(shares, f)

@app.route("/api/compass-share", methods=["POST"])
@login_required
def create_compass_share():
    data = request.get_json()
    job = (data or {}).get("job", "").strip()
    if not job:
        return jsonify({"error": "Missing job"}), 400
    shares = load_compass_shares()
    shares = {t: v for t, v in shares.items() if v.get("job") != job}
    token = secrets.token_urlsafe(16)
    shares[token] = {"job": job, "created": datetime.utcnow().isoformat()}
    save_compass_shares(shares)
    return jsonify({"token": token, "url": f"/compass/{token}"})

@app.route("/api/compass-share/info", methods=["POST"])
@login_required
def compass_share_info():
    data = request.get_json()
    job = (data or {}).get("job", "").strip()
    shares = load_compass_shares()
    for token, v in shares.items():
        if v.get("job") == job:
            return jsonify({"token": token, "url": f"/compass/{token}"})
    return jsonify({"token": None})

@app.route("/api/compass-share/delete", methods=["POST"])
@login_required
def delete_compass_share():
    data = request.get_json()
    token = (data or {}).get("token", "").strip()
    shares = load_compass_shares()
    shares.pop(token, None)
    save_compass_shares(shares)
    return jsonify({"ok": True})

@app.route("/compass/<token>")
def view_compass_share(token):
    shares = load_compass_shares()
    info = shares.get(token)
    if not info:
        return "This link has expired or does not exist.", 404
    job = info["job"]
    bp_path = os.path.join(JOBS_DIR, job, "Blueprints")
    blueprints = sorted([f for f in os.listdir(bp_path) if f.lower().endswith(".pdf")]) if os.path.isdir(bp_path) else []
    return render_template("compass_share.html", token=token, job=job, blueprints=blueprints)

@app.route("/compass/<token>/file/<path:filepath>")
def serve_compass_share_file(token, filepath):
    shares = load_compass_shares()
    info = shares.get(token)
    if not info:
        return "Link not found", 404
    full = safe_join(info["job"], "Blueprints", filepath)
    return send_from_directory(os.path.dirname(full), os.path.basename(full))

# ── Blueprint Hyperlinks ──────────────────────────────────────
import threading, shutil, uuid as _uuid

# In-memory job store: job_id -> {status, result_path, out_name, added, error}
_hl_jobs = {}
_hl_lock = threading.Lock()


def _find_d_pages(doc):
    """
    Scan every page for D-page titles using font size as the signal.
    For each D-number (D1–D30), the page where it appears in the LARGEST
    font is the title page for that detail — works on any KPS blueprint.
    Returns dict: {dn: page_index}
    """
    import fitz
    # d_best: dn -> (page_index, max_font_size_seen)
    d_best = {}
    for pi in range(len(doc)):
        page = doc[pi]
        try:
            blocks = page.get_text("dict")["blocks"]
        except Exception:
            continue
        for block in blocks:
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = span.get("text", "")
                    sz  = span.get("size", 0)
                    for m in re.finditer(r'\bD(\d{1,2})\b', txt):
                        dn = int(m.group(1))
                        if 1 <= dn <= 30:
                            if dn not in d_best or sz > d_best[dn][1]:
                                d_best[dn] = (pi, sz)
    return {dn: pi for dn, (pi, _) in d_best.items()}


def _run_hyperlink_job(job_id, pdf_path, display_name):
    """Background thread: detect callout circles, find D-pages, insert links."""
    import fitz
    from callout_engine import detect_callouts_on_page

    out_path = pdf_path + "_out.pdf"
    try:
        doc = fitz.open(pdf_path)

        # 1. Detect callouts on every page
        all_callouts = []
        for pi in range(len(doc)):
            all_callouts.extend(detect_callouts_on_page(doc, pi))

        # 2. Find D-pages by largest font occurrence of "Dn"
        needed_dns = {c["dp"] for c in all_callouts}
        all_d_pages = _find_d_pages(doc)
        d_page_map = {dn: pi for dn, pi in all_d_pages.items() if dn in needed_dns}

        # 3. Apply orange highlights + GoTo links (link to top of D-page)
        ORANGE, BORDER = (1.0, 0.65, 0.2), (0.85, 0.45, 0.0)
        LINK_SIZE = 49  # match manual stamp size (px in PDF points)
        added = 0
        for c in all_callouts:
            dest_pi = d_page_map.get(c["dp"])
            if dest_pi is None:
                continue
            # Centre the fixed-size rect on the detected callout centre
            cx, cy = c["cx"], c["cy"]
            half = LINK_SIZE / 2
            rect = fitz.Rect(cx - half, cy - half, cx + half, cy + half)
            page = doc[c["pi"]]
            sh = page.new_shape()
            sh.draw_rect(rect)
            sh.finish(color=BORDER, fill=ORANGE, fill_opacity=0.35, width=1.2)
            sh.commit()
            page.insert_link({"kind": fitz.LINK_GOTO, "from": rect,
                               "page": dest_pi, "to": fitz.Point(0, 0)})
            added += 1

        doc.save(out_path, garbage=4, deflate=True, incremental=False)
        doc.close()

        out_name = re.sub(r'\.pdf$', '', display_name, flags=re.IGNORECASE) + "_linked.pdf"
        with _hl_lock:
            _hl_jobs[job_id] = {"status": "done", "result_path": out_path,
                                 "out_name": out_name, "added": added,
                                 "d_page_map": all_d_pages}
    except Exception as e:
        with _hl_lock:
            _hl_jobs[job_id] = {"status": "error", "error": f"{type(e).__name__}: {e}"}
    finally:
        try:
            os.unlink(pdf_path)
        except Exception:
            pass


@app.route("/api/blueprint-hyperlinks/files")
@login_required
def api_blueprint_hyperlink_files():
    """Return all Blueprint PDFs grouped by job."""
    if not os.path.isdir(JOBS_DIR):
        return jsonify([])
    result = []
    for job in sorted(os.listdir(JOBS_DIR)):
        job_path = os.path.join(JOBS_DIR, job)
        if not os.path.isdir(job_path):
            continue
        bp_path = os.path.join(job_path, "Blueprints")
        if not os.path.isdir(bp_path):
            continue
        for f in sorted(os.listdir(bp_path)):
            if f.lower().endswith(".pdf"):
                result.append({"job": job, "filename": f})
    return jsonify(result)


@app.route("/blueprint/hyperlinks", methods=["GET", "POST"])
@login_required
def blueprint_hyperlinks():
    if request.method == "GET":
        return render_template("blueprint_hyperlinks.html")

    job_name = request.form.get("job", "").strip()
    job_file = request.form.get("filename", "").strip()
    uploaded = request.files.get("pdf")

    # Write PDF to a temp file for the background thread
    tmp_in = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    try:
        if job_name and job_file:
            server_path = safe_join(job_name, "Blueprints", job_file)
            if not os.path.isfile(server_path):
                return jsonify({"error": f"File not found: {job_file}"}), 404
            tmp_in.close()
            shutil.copy2(server_path, tmp_in.name)
            display_name = job_file
        elif uploaded and uploaded.filename.lower().endswith(".pdf"):
            uploaded.save(tmp_in)
            tmp_in.close()
            display_name = uploaded.filename
        else:
            tmp_in.close()
            os.unlink(tmp_in.name)
            return jsonify({"error": "Please select a blueprint or upload a PDF."}), 400
    except Exception as e:
        try: os.unlink(tmp_in.name)
        except: pass
        return jsonify({"error": str(e)}), 500

    job_id = str(_uuid.uuid4())
    with _hl_lock:
        _hl_jobs[job_id] = {
            "status": "processing",
            "source_job": job_name if job_name else None,
            "source_filename": job_file if job_file else None,
        }

    t = threading.Thread(target=_run_hyperlink_job,
                         args=(job_id, tmp_in.name, display_name), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/blueprint/hyperlinks/open", methods=["POST"])
@login_required
def blueprint_hyperlinks_open():
    """Open a PDF directly in the editor without running OCR processing."""
    job_name = request.form.get("job", "").strip()
    job_file = request.form.get("filename", "").strip()
    uploaded = request.files.get("pdf")

    tmp_in = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    try:
        if job_name and job_file:
            server_path = safe_join(job_name, "Blueprints", job_file)
            if not os.path.isfile(server_path):
                return jsonify({"error": f"File not found: {job_file}"}), 404
            tmp_in.close()
            shutil.copy2(server_path, tmp_in.name)
            display_name = job_file
        elif uploaded and uploaded.filename.lower().endswith(".pdf"):
            uploaded.save(tmp_in)
            tmp_in.close()
            display_name = uploaded.filename
        else:
            tmp_in.close()
            os.unlink(tmp_in.name)
            return jsonify({"error": "Please select a blueprint or upload a PDF."}), 400
    except Exception as e:
        try: os.unlink(tmp_in.name)
        except: pass
        return jsonify({"error": str(e)}), 500

    # Find D-pages so the editor knows about D-labels
    try:
        import fitz as _fitz
        doc = _fitz.open(tmp_in.name)
        d_page_map = _find_d_pages(doc)
        doc.close()
    except Exception:
        d_page_map = {}

    out_name = re.sub(r'\.pdf$', '', display_name, flags=re.IGNORECASE) + "_edited.pdf"
    job_id = str(_uuid.uuid4())
    with _hl_lock:
        _hl_jobs[job_id] = {
            "status": "done",
            "result_path": tmp_in.name,
            "out_name": out_name,
            "added": 0,
            "d_page_map": d_page_map,
            "source_job": job_name if job_name else None,
            "source_filename": job_file if job_file else None,
        }
    return jsonify({"job_id": job_id})


@app.route("/blueprint/hyperlinks/status/<job_id>")
@login_required
def blueprint_hyperlinks_status(job_id):
    with _hl_lock:
        job = _hl_jobs.get(job_id)
    if not job:
        return jsonify({"status": "not_found"}), 404
    return jsonify({k: v for k, v in job.items() if k != "result_path"})


@app.route("/blueprint/hyperlinks/download/<job_id>")
@login_required
def blueprint_hyperlinks_download(job_id):
    with _hl_lock:
        job = _hl_jobs.get(job_id)
    if not job or job.get("status") != "done":
        return "Not ready", 404
    return send_file(job["result_path"], mimetype="application/pdf",
                     as_attachment=True, download_name=job["out_name"])


@app.route("/blueprint/hyperlinks/editor/<job_id>")
@login_required
def blueprint_hyperlinks_editor(job_id):
    with _hl_lock:
        job = _hl_jobs.get(job_id)
    if not job or job.get("status") != "done":
        return "Job not found or not ready.", 404
    return render_template("blueprint_hyperlinks_editor.html",
                           job_id=job_id,
                           out_name=job["out_name"],
                           source_job=job.get("source_job") or "",
                           source_filename=job.get("source_filename") or "")


@app.route("/blueprint/hyperlinks/view/<job_id>")
@login_required
def blueprint_hyperlinks_view(job_id):
    """Serve the processed PDF for the editor viewer."""
    with _hl_lock:
        job = _hl_jobs.get(job_id)
    if not job or job.get("status") != "done":
        return "Not ready", 404
    return send_file(job["result_path"], mimetype="application/pdf")


@app.route("/blueprint/hyperlinks/links/<job_id>")
@login_required
def blueprint_hyperlinks_links(job_id):
    """Return all GoTo links from the processed PDF as JSON."""
    import fitz as _fitz
    with _hl_lock:
        job = _hl_jobs.get(job_id)
    if not job or job.get("status") != "done":
        return jsonify({"error": "Not ready"}), 404
    try:
        doc = _fitz.open(job["result_path"])
        pages_info = []
        links = []
        for pi in range(len(doc)):
            page = doc[pi]
            pages_info.append({"width": page.rect.width, "height": page.rect.height})
            for lk in page.get_links():
                if lk.get("kind") == _fitz.LINK_GOTO:
                    r = lk["from"]
                    links.append({
                        "page": pi,
                        "rect": [r.x0, r.y0, r.x1, r.y1],
                        "dest_page": lk.get("page", 0)
                    })
        doc.close()
        d_page_map = {str(k): v for k, v in (job.get("d_page_map") or {}).items()}
        return jsonify({"pages": pages_info, "links": links, "d_pages": d_page_map})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/blueprint/hyperlinks/save/<job_id>", methods=["POST"])
@login_required
def blueprint_hyperlinks_save(job_id):
    """Accept edited links JSON, rebuild PDF, return as download."""
    import fitz as _fitz
    with _hl_lock:
        job = _hl_jobs.get(job_id)
    if not job or job.get("status") != "done":
        return jsonify({"error": "Job not found"}), 404

    data = request.get_json()
    edited_links = data.get("links", [])

    try:
        doc = _fitz.open(job["result_path"])

        # Remove all existing GoTo links
        for pi in range(len(doc)):
            page = doc[pi]
            for lk in page.get_links():
                if lk.get("kind") == _fitz.LINK_GOTO:
                    page.delete_link(lk)

        # Re-insert links from edited data (orange highlight already in PDF)
        for lk in edited_links:
            pi = lk["page"]
            if pi < 0 or pi >= len(doc):
                continue
            dest = lk["dest_page"]
            if dest < 0 or dest >= len(doc):
                continue
            rect = _fitz.Rect(lk["rect"])
            doc[pi].insert_link({
                "kind": _fitz.LINK_GOTO,
                "from": rect,
                "page": dest,
                "to": _fitz.Point(0, 0)
            })

        buf = io.BytesIO()
        doc.save(buf, garbage=4, deflate=True, incremental=False)
        doc.close()
        buf.seek(0)

        # Update stored file with saved version
        with open(job["result_path"], "wb") as f:
            f.write(buf.getvalue())
        buf.seek(0)

        return send_file(buf, mimetype="application/pdf",
                         as_attachment=True, download_name=job["out_name"])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/blueprint/hyperlinks/publish/<job_id>", methods=["POST"])
@login_required
def blueprint_hyperlinks_publish(job_id):
    """Replace the original blueprint on the server with the linked version.
    Archives the old file in Old Versions/ and deletes archives older than 30 days."""
    import fitz as _fitz
    from datetime import datetime, timedelta

    with _hl_lock:
        job = _hl_jobs.get(job_id)
    if not job or job.get("status") != "done":
        return jsonify({"error": "Job not found"}), 404

    source_job      = job.get("source_job")
    source_filename = job.get("source_filename")
    if not source_job or not source_filename:
        return jsonify({"error": "No source file to publish to — this PDF was uploaded, not selected from a job."}), 400

    # Paths
    bp_dir      = safe_join(source_job, "Blueprints")
    target_path = os.path.join(bp_dir, source_filename)
    old_dir     = os.path.join(bp_dir, "Old Versions")
    os.makedirs(old_dir, exist_ok=True)

    # Save current editor state to the result file first
    data = request.get_json(silent=True) or {}
    edited_links = data.get("links")
    if edited_links is not None:
        try:
            doc = _fitz.open(job["result_path"])
            for pi in range(len(doc)):
                for lk in doc[pi].get_links():
                    if lk.get("kind") == _fitz.LINK_GOTO:
                        doc[pi].delete_link(lk)
            for lk in edited_links:
                pi = lk["page"]
                if 0 <= pi < len(doc) and 0 <= lk["dest_page"] < len(doc):
                    doc[pi].insert_link({
                        "kind": _fitz.LINK_GOTO,
                        "from": _fitz.Rect(lk["rect"]),
                        "page": lk["dest_page"],
                        "to": _fitz.Point(0, 0)
                    })
            doc.save(job["result_path"] + "_pub.pdf", garbage=4, deflate=True, incremental=False)
            doc.close()
            pub_path = job["result_path"] + "_pub.pdf"
        except Exception as e:
            return jsonify({"error": f"Failed to save links: {e}"}), 500
    else:
        pub_path = job["result_path"]

    try:
        # Archive old file with timestamp
        if os.path.isfile(target_path):
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            base = os.path.splitext(source_filename)[0]
            archive_name = f"{base}__{ts}.pdf"
            shutil.move(target_path, os.path.join(old_dir, archive_name))

        # Copy new file into place
        shutil.copy2(pub_path, target_path)

        # Clean up tmp pub file
        if pub_path != job["result_path"]:
            try: os.unlink(pub_path)
            except: pass

        # Delete Old Versions files older than 30 days
        cutoff = datetime.utcnow() - timedelta(days=30)
        for fname in os.listdir(old_dir):
            fpath = os.path.join(old_dir, fname)
            if os.path.isfile(fpath):
                mtime = datetime.utcfromtimestamp(os.path.getmtime(fpath))
                if mtime < cutoff:
                    try: os.unlink(fpath)
                    except: pass

        return jsonify({"ok": True, "message": f"Published to {source_job} / {source_filename}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500






# ── Packing List Tracker ──────────────────────────────────────────────────────
import threading as _threading
import json as _json

_pl_jobs      = {}
_pl_jobs_lock = _threading.Lock()
# Per-job file locks — serialise delivery_state read-modify-write so processing
# multiple packing lists simultaneously doesn't cause threads to clobber each other.
_pl_file_locks      = {}
_pl_file_locks_meta = _threading.Lock()

def _pl_get_file_lock(job_name):
    with _pl_file_locks_meta:
        if job_name not in _pl_file_locks:
            _pl_file_locks[job_name] = _threading.Lock()
        return _pl_file_locks[job_name]

def _pl_tracking_dir(job_name):  return safe_join(job_name, "Delivery Tracking")
def _pl_dxf_dir(job_name):
    d = safe_join(job_name, "DXF CAD FILE")
    return d if os.path.isdir(d) else None
def _pl_state_path(job_name):    return os.path.join(_pl_tracking_dir(job_name), "delivery_state.json")
def _pl_cache_path(job_name):    return os.path.join(_pl_tracking_dir(job_name), "panel_locations_v2.json")  # v2 = DXF-coordinate locator
def _pl_output_path(job_name):   return os.path.join(_pl_tracking_dir(job_name), "tracked_blueprint.pdf")
def _pl_cells_path(job_name):    return os.path.join(_pl_tracking_dir(job_name), "table_cells.json")
def _pl_colors_path(job_name):    return os.path.join(_pl_tracking_dir(job_name), "ship_colors.json")
def _pl_registry_path(job_name): return os.path.join(_pl_tracking_dir(job_name), "shipment_registry.json")

# Persistent shipment → color-index map. A color is assigned to a packing list the
# first time it's processed and never changes; the SAME index is used by the tracker
# UI, the editor, and the baked prints. (Palettes have 7 colors, no red.)
def _pl_load_colors(job_name):
    p = _pl_colors_path(job_name)
    if os.path.exists(p):
        try:
            with open(p) as f:
                return _json.load(f)
        except Exception:
            pass
    return {}

def _pl_assign_colors(job_name, delivery_state, extra_shipments=None):
    """Ensure every shipment present in delivery_state has a stored color index.
    New shipments get the lowest unused index (in first-seen order). Returns the map.
    extra_shipments: list of labels to ensure are in the map even if not in delivery_state."""
    m = _pl_load_colors(job_name)
    new = []
    for info in delivery_state.values():
        s = info.get("shipment", "")
        if s and s not in m and s not in new:
            new.append(s)
    for s in (extra_shipments or []):
        if s and s not in m and s not in new:
            new.append(s)
    for s in new:
        used = set(m.values())
        i = 0
        while i in used:
            i += 1
        m[s] = i
    try:
        os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
        with open(_pl_colors_path(job_name), "w") as f:
            _json.dump(m, f)
    except Exception:
        pass
    return m

def _pl_record_shipment(job_name, label, parsed_skids, cmap):
    """Persist every processed shipment label + skids to the registry, even if 0 panels.
    This ensures 0-panel shipments still appear in section 3 of the tracker UI."""
    reg_path = _pl_registry_path(job_name)
    try:
        reg = {}
        if os.path.exists(reg_path):
            with open(reg_path) as f:
                reg = _json.load(f)
        _sk = lambda x: (int(x) if str(x).isdigit() else float('inf'), str(x))
        reg[label] = {
            "skids": sorted([str(s) for s in parsed_skids], key=_sk),
            "color_index": cmap.get(label, 0),
        }
        with open(reg_path, "w") as f:
            _json.dump(reg, f)
    except Exception:
        pass

def _find_blueprint(job_name):
    bp_dir = safe_join(job_name, "Blueprints")
    if os.path.isdir(bp_dir):
        for f in sorted(os.listdir(bp_dir)):
            if f.lower().endswith(".pdf") and "Old Versions" not in f and "Delivery Tracked" not in f:
                return os.path.join(bp_dir, f)
    return None

def _run_pl_job(job_name, packing_list_path, shipment_label):
    try:
        from packing_list_engine import (parse_packing_list, scan_blueprint_panels,
                                         generate_tracked_blueprint)
        os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)

        with _pl_jobs_lock:
            _pl_jobs[job_name].update({"message": "Parsing packing list...", "progress": 10})

        panels_added = 0
        parsed, parse_warnings = parse_packing_list(packing_list_path)
        with _pl_jobs_lock:
            skid_count = len(parsed)
            raw_panels = sum(len(v) for v in parsed.values())
            _pl_jobs[job_name].update({"message": f"Parsed {raw_panels} panels across {skid_count} skids…", "progress": 15})

        # Serialise state read-modify-write: if multiple packing lists are processed
        # simultaneously each thread must hold the per-job lock while touching the
        # delivery_state file or it will clobber the other threads' panels.
        _file_lock = _pl_get_file_lock(job_name)
        panels_reassigned = 0
        with _file_lock:
            delivery_state = {}
            if os.path.exists(_pl_state_path(job_name)):
                with open(_pl_state_path(job_name)) as f:
                    delivery_state = _json.load(f)

            for skid_num, panel_orders in parsed.items():
                for p, order_num in panel_orders.items():
                    if p not in delivery_state:
                        panels_added += 1
                    elif delivery_state[p].get("shipment") != shipment_label:
                        # Panel exists from a prior shipment — reassign to this one so
                        # the blueprint markup color matches the tracker's color for this shipment.
                        panels_reassigned += 1
                    delivery_state[p] = {"skid": skid_num, "shipment": shipment_label,
                                         "order_num": order_num}

            os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
            with open(_pl_state_path(job_name), "w") as f:
                _json.dump(delivery_state, f)

        with _pl_jobs_lock:
            msg = f"Added {panels_added} new panels"
            if panels_reassigned:
                msg += f", reassigned {panels_reassigned} to this shipment"
            msg += " — scanning blueprint…"
            _pl_jobs[job_name].update({"message": msg, "progress": 18})

        # ── Use Panel Mapper session if one exists ───────────────────────────
        # The Panel Mapper has already located every panel precisely.  Prefer
        # those positions over an OCR scan; use the mapper's scan PDF as the
        # base so the output shows the clean panel-layout pages.
        pm_sess      = _pm_load_session(job_name)
        pm_locs_path = pm_sess.get("locs", "") if pm_sess else ""
        pm_scan_pdf  = pm_sess.get("scan_pdf", "") if pm_sess else ""
        use_panel_map = (pm_sess and
                         pm_locs_path and os.path.isfile(pm_locs_path) and
                         pm_scan_pdf  and os.path.isfile(pm_scan_pdf))

        if use_panel_map:
            with open(pm_locs_path) as _f:
                panel_locations = _json.load(_f)
            # Write to the tracker's cache so the editor right-pane also uses
            # these positions (coloring packing-list marks by panel presence).
            try:
                os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
                with open(_pl_cache_path(job_name), "w") as _f:
                    _json.dump(panel_locations, _f)
            except Exception:
                pass
            with _pl_jobs_lock:
                _pl_jobs[job_name].update({
                    "message": f"Using Panel Mapper positions ({len(panel_locations)} panels located)…",
                    "progress": 50
                })
        else:
            # Fall back to OCR + optional DXF validation
            blueprint_path = _find_blueprint(job_name)
            if not blueprint_path:
                raise FileNotFoundError("No blueprint PDF found for this job.")

            def progress_cb(pg, total):
                with _pl_jobs_lock:
                    _pl_jobs[job_name].update({
                        "progress": 20 + int(60 * pg / max(total, 1)),
                        "message": f"Scanning blueprint page {pg+1}/{total}..."
                    })

            dxf_dir = _pl_dxf_dir(job_name)
            if dxf_dir:
                try:
                    import ezdxf as _ezdxf_check
                    dxf_status = f"DXF validation active"
                except ImportError:
                    dxf_status = "⚠ ezdxf not installed (run deploy.bat)"
                    dxf_dir = None
            else:
                dxf_status = "⚠ No DXF folder — panel numbers not validated"
            with _pl_jobs_lock:
                _pl_jobs[job_name].update({"message": f"Scanning blueprint… {dxf_status}", "progress": 19})

            panel_locations = scan_blueprint_panels(
                blueprint_path, _pl_cache_path(job_name), progress_cb, dxf_dir=dxf_dir)

        # Re-read delivery_state under the lock so we incorporate any panels that
        # sibling threads wrote while we were scanning the blueprint.
        with _file_lock:
            if os.path.exists(_pl_state_path(job_name)):
                with open(_pl_state_path(job_name)) as f:
                    delivery_state = _json.load(f)

            corr = _pl_load_corrections(job_name)
            if any(corr.get(k) for k in ("deletions", "renames", "additions")):
                _apply_corrections(delivery_state, panel_locations, corr)
                with open(_pl_state_path(job_name), "w") as f:
                    _json.dump(delivery_state, f)
                if not use_panel_map:
                    with open(_pl_cache_path(job_name), "w") as f:
                        _json.dump(panel_locations, f)

            # Lock in color for this shipment even if it has 0 panels
            ship_colors = _pl_assign_colors(job_name, delivery_state, extra_shipments=[shipment_label])
            _pl_record_shipment(job_name, shipment_label, list(parsed.keys()), ship_colors)

        with _pl_jobs_lock:
            _pl_jobs[job_name].update({"progress": 85, "message": "Generating annotated blueprint..."})
        if use_panel_map:
            from packing_list_engine import generate_tracked_blueprint_panel_map
            generate_tracked_blueprint_panel_map(
                pm_scan_pdf, panel_locations, delivery_state, _pl_output_path(job_name),
                shipment_colors=ship_colors)
            table_cells = {}
        else:
            table_cells = generate_tracked_blueprint(
                blueprint_path, delivery_state, panel_locations, _pl_output_path(job_name),
                shipment_colors=ship_colors)
        try:
            with open(_pl_cells_path(job_name), "w") as f:
                _json.dump(table_cells or {}, f)
        except Exception:
            pass

        located = sum(1 for p in delivery_state if p in panel_locations)
        done_msg = f"Complete — {panels_added} new panels added"
        if panels_reassigned:
            done_msg += f", {panels_reassigned} reassigned to this shipment"
        done_msg += f", {located} located on blueprint"
        with _pl_jobs_lock:
            _pl_jobs[job_name].update({
                "status": "done", "progress": 100,
                "message": done_msg,
                "panels_added": panels_added, "total_panels": len(delivery_state),
                "warnings": parse_warnings,
                "located": located,
                "skids": sorted(set(v["skid"] for v in delivery_state.values()), key=lambda x: (int(x) if str(x).isdigit() else float('inf'), str(x))),
            })
    except Exception as e:
        with _pl_jobs_lock:
            _pl_jobs[job_name].update({"status": "error", "message": str(e)})


# ── Panel Print Mapper ────────────────────────────────────────────────────────
# Verification tool: render a job's blueprint with every panel boxed in red and
# its number labeled above, so a human can confirm panels are read correctly.

_pm_jobs = {}                     # job_name -> {status, progress, message, ...}
_pm_jobs_lock = _threading.Lock()

def _pm_dir(job_name):
    return safe_join(job_name, "Panel Map")

def _pm_safe(name):
    base = os.path.splitext(os.path.basename(name))[0]
    return re.sub(r"[^A-Za-z0-9._-]+", "_", base)[:80]

def _pm_base(bp_name):
    """Canonical blueprint name. Re-loading the full 'Panel Mapper' version maps
    back to the same working set as the original (continues progress); the
    'panels_only' subset stays its own set so saving it can't overwrite the full."""
    b = os.path.splitext(os.path.basename(bp_name))[0]
    if b.endswith(" - Panel Mapper"):
        b = b[:-len(" - Panel Mapper")]
    return b.strip()

def _pm_sig(pages):
    """Short signature for a kept-pages selection (None/empty = whole doc)."""
    if not pages:
        return ""
    import hashlib
    key = ",".join(str(p) for p in sorted(pages))
    return "_p" + hashlib.md5(key.encode()).hexdigest()[:8]

def _pm_cache_path(job_name, bp_name, pages=None):
    return os.path.join(_pm_dir(job_name), f"locs_{_pm_safe(_pm_base(bp_name))}{_pm_sig(pages)}.json")

def _pm_output_path(job_name, bp_name):
    return os.path.join(_pm_dir(job_name), f"map_{_pm_safe(_pm_base(bp_name))}.pdf")   # marked-up pages only

def _pm_full_path(job_name, bp_name):
    return os.path.join(_pm_dir(job_name), f"full_{_pm_safe(_pm_base(bp_name))}.pdf")  # whole doc w/ edits merged

def _pm_merge_full(original_path, map_path, pages, dest):
    """Write `dest` = the complete original blueprint with the marked-up pages
    merged back in at their original positions. `pages` = 1-based originals."""
    import fitz, shutil
    if not pages:
        shutil.copy2(map_path, dest)   # whole doc was mapped already
        return
    full = fitz.open(original_path)
    mapped = fitz.open(map_path)
    try:
        n = full.page_count
        for i, op in enumerate(sorted(pages)):
            idx = op - 1
            if idx < 0 or idx >= n or i >= mapped.page_count:
                continue
            full.delete_page(idx)                                  # drop original page
            full.insert_pdf(mapped, from_page=i, to_page=i, start_at=idx)  # put marked-up one back
        full.save(dest, garbage=4, deflate=True)
    finally:
        full.close(); mapped.close()

def _pm_build_full(job_name, bp_name, pages, src_pdf):
    """Regenerate the full document by merging the marked-up pages back into the
    source document the user is working on (original blueprint or a saved one)."""
    map_out = _pm_output_path(job_name, bp_name)
    if not src_pdf or not os.path.isfile(src_pdf) or not os.path.isfile(map_out):
        return
    try:
        _pm_merge_full(src_pdf, map_out, pages, _pm_full_path(job_name, bp_name))
    except Exception:
        pass

def _pm_publish_to_viewer(job_name, bp_name):
    """Copy the current panel-mapper documents into the job's 'Panel Mapper'
    folder so the Blueprint Viewer shows them. Stable names = updated each save."""
    import shutil
    try:
        dest_dir = safe_join(job_name, "Panel Mapper")
        os.makedirs(dest_dir, exist_ok=True)
        base   = _pm_base(bp_name)
        full   = _pm_full_path(job_name, bp_name)
        mapped = _pm_output_path(job_name, bp_name)
        if os.path.isfile(full):
            shutil.copy2(full, os.path.join(dest_dir, f"{base} - Panel Mapper.pdf"))
        if os.path.isfile(mapped):
            shutil.copy2(mapped, os.path.join(dest_dir, f"{base} - panels_only.pdf"))
    except Exception:
        pass

def _pm_session_path(job_name):
    return os.path.join(_pm_dir(job_name), "session.json")

def _pm_load_session(job_name):
    p = _pm_session_path(job_name)
    if os.path.exists(p):
        try:
            with open(p) as f:
                return _json.load(f)
        except Exception:
            pass
    return None

def _pm_page_dims(pdf_path):
    import fitz
    doc = fitz.open(pdf_path)
    dims = [{"w": pg.rect.width, "h": pg.rect.height} for pg in doc]
    doc.close()
    return dims

def _pm_make_trimmed(bp_path, pages, dest):
    """Write a new PDF containing only `pages` (1-based) from bp_path, in order."""
    import fitz
    src = fitz.open(bp_path)
    keep = sorted({p - 1 for p in pages if 1 <= p <= src.page_count})
    if not keep:
        keep = list(range(src.page_count))
    src.select(keep)
    src.save(dest, garbage=4, deflate=True)
    src.close()
    return len(keep)

def _run_pm_job(job_name, bp_name, pages=None, folder="Blueprints"):
    try:
        from packing_list_engine import scan_blueprint_panels, generate_panel_map_blueprint
        os.makedirs(_pm_dir(job_name), exist_ok=True)

        if folder not in _PM_SRC_FOLDERS:
            folder = "Blueprints"
        bp_path = safe_join(job_name, folder, bp_name)
        if not os.path.isfile(bp_path):
            raise FileNotFoundError(f"Document not found: {bp_name}")

        # If the user pre-selected pages, scan a trimmed copy so blank pages are
        # never even OCR'd (much faster). Otherwise scan the whole blueprint.
        scan_path = bp_path
        if pages:
            trimmed = os.path.join(_pm_dir(job_name), f"trimmed_{_pm_safe(_pm_base(bp_name))}{_pm_sig(pages)}.pdf")
            kept_n = _pm_make_trimmed(bp_path, pages, trimmed)
            scan_path = trimmed
            with _pm_jobs_lock:
                _pm_jobs[job_name].update({"message": f"Scanning {kept_n} selected pages…"})

        dxf_dir = _pl_dxf_dir(job_name)
        if dxf_dir:
            try:
                import ezdxf as _ezdxf_check  # noqa: F401
                dxf_status = "DXF validation active"
            except ImportError:
                dxf_status = "⚠ ezdxf not installed — panels not validated"
                dxf_dir = None
        else:
            dxf_status = "⚠ No DXF folder — panel numbers not validated"

        with _pm_jobs_lock:
            _pm_jobs[job_name].update({"message": f"Scanning blueprint… {dxf_status}", "progress": 10})

        def progress_cb(pg, total):
            with _pm_jobs_lock:
                _pm_jobs[job_name].update({
                    "progress": 10 + int(70 * pg / max(total, 1)),
                    "message": f"Scanning page {pg+1}/{total}…",
                })

        panel_locations = scan_blueprint_panels(
            scan_path, _pm_cache_path(job_name, bp_name, pages), progress_cb, dxf_dir=dxf_dir)

        with _pm_jobs_lock:
            _pm_jobs[job_name].update({"progress": 85, "message": "Drawing panel map…"})

        # Keep ALL loaded pages in the output (the user already chose which pages
        # to include), even pages where no panels were detected.
        result = generate_panel_map_blueprint(
            scan_path, panel_locations, _pm_output_path(job_name, bp_name),
            keep_only_panel_pages=False)
        _pm_build_full(job_name, bp_name, pages, bp_path)   # merge edits back into the source
        _pm_publish_to_viewer(job_name, bp_name)
        drawn   = result["drawn"]
        out_pgs = result["output_pages"]
        withp   = result["pages_with_panels"]

        # Save a session so the editor can reload the base prints + panel positions.
        try:
            with open(_pm_session_path(job_name), "w") as f:
                _json.dump({"bp_name": bp_name, "folder": folder, "src_pdf": bp_path,
                            "scan_pdf": scan_path,
                            "locs": _pm_cache_path(job_name, bp_name, pages),
                            "pages": pages or []}, f)
        except Exception:
            pass

        with _pm_jobs_lock:
            _pm_jobs[job_name].update({
                "status": "done", "progress": 100, "blueprint": bp_name,
                "message": f"Complete — {drawn} panels mapped over {out_pgs} pages "
                           f"({withp} page(s) had panels).",
                "panels": drawn, "output_pages": out_pgs,
                "pages_with_panels": withp, "panel_pages": result["panel_pages"],
            })
    except Exception as e:
        with _pm_jobs_lock:
            _pm_jobs[job_name].update({"status": "error", "message": str(e)})


@app.route("/panel-print-mapper")
@login_required
def panel_print_mapper():
    return render_template("panel_print_mapper.html")


@app.route("/api/panel-map/blueprints/<path:job_name>")
@login_required
def panel_map_blueprints(job_name):
    """List source PDFs: the original blueprints plus any saved Panel Mapper
    documents (so prior progress can be re-loaded). If there's no Panel Mapper
    folder, only the originals are returned."""
    def _list(folder, skip_tracked=False):
        out = []
        d0 = safe_join(job_name, folder)
        if os.path.isdir(d0):
            for f in sorted(os.listdir(d0)):
                if not f.lower().endswith(".pdf"):
                    continue
                if skip_tracked and "Delivery Tracked" in f:
                    continue
                try:
                    dt = datetime.fromtimestamp(os.path.getmtime(os.path.join(d0, f)))
                    date = f"{dt.strftime('%b')} {dt.day}, {dt.year}"
                except Exception:
                    date = ""
                out.append({"name": f, "date": date, "folder": folder})
        return out
    files = _list("Blueprints", skip_tracked=True) + _list("Panel Mapper")
    return jsonify({"blueprints": files, "has_dxf": _pl_dxf_dir(job_name) is not None,
                    "has_session": os.path.isfile(_pm_session_path(job_name))})


# Folders the panel mapper is allowed to read source documents from.
_PM_SRC_FOLDERS = ("Blueprints", "Panel Mapper")


@app.route("/api/panel-map/process/<path:job_name>", methods=["POST"])
@login_required
def panel_map_process(job_name):
    data    = request.get_json(silent=True) or {}
    bp_name = (data.get("blueprint") or "").strip()
    folder  = data.get("folder") if data.get("folder") in _PM_SRC_FOLDERS else "Blueprints"
    if not bp_name or not bp_name.lower().endswith(".pdf"):
        return jsonify({"error": "Choose a blueprint PDF"}), 400
    if not os.path.isfile(safe_join(job_name, folder, bp_name)):
        return jsonify({"error": f"Document not found: {bp_name}"}), 404

    # Optional manual page selection (1-based page numbers to keep / scan).
    pages = data.get("pages")
    if pages:
        try:
            pages = sorted({int(p) for p in pages if int(p) >= 1})
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid page selection"}), 400
        if not pages:
            return jsonify({"error": "Select at least one page"}), 400
    else:
        pages = None

    if data.get("rescan"):
        cp = _pm_cache_path(job_name, bp_name, pages)
        if os.path.exists(cp):
            os.unlink(cp)
    with _pm_jobs_lock:
        _pm_jobs[job_name] = {"status": "processing", "progress": 0,
                              "message": "Starting…", "blueprint": bp_name}
    _threading.Thread(target=_run_pm_job, args=(job_name, bp_name, pages, folder), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/panel-map/load-only/<path:job_name>", methods=["POST"])
@login_required
def panel_map_load_only(job_name):
    """Load the selected pages into the editor WITHOUT detecting panels — the user
    will place every panel by hand. Sets up an empty session and opens the editor."""
    from packing_list_engine import generate_panel_map_blueprint
    data    = request.get_json(silent=True) or {}
    bp_name = (data.get("blueprint") or "").strip()
    folder  = data.get("folder") if data.get("folder") in _PM_SRC_FOLDERS else "Blueprints"
    if not bp_name or not bp_name.lower().endswith(".pdf"):
        return jsonify({"error": "Choose a blueprint PDF"}), 400
    bp_path = safe_join(job_name, folder, bp_name)
    if not os.path.isfile(bp_path):
        return jsonify({"error": f"Document not found: {bp_name}"}), 404

    pages = data.get("pages")
    if pages:
        try:
            pages = sorted({int(p) for p in pages if int(p) >= 1})
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid page selection"}), 400
        if not pages:
            return jsonify({"error": "Select at least one page"}), 400
    else:
        pages = None

    os.makedirs(_pm_dir(job_name), exist_ok=True)

    # Trim to the selected pages (or use the whole document).
    scan_path = bp_path
    if pages:
        scan_path = os.path.join(_pm_dir(job_name),
                                 f"trimmed_{_pm_safe(_pm_base(bp_name))}{_pm_sig(pages)}.pdf")
        _pm_make_trimmed(bp_path, pages, scan_path)

    # Empty panel set — the editor starts with a clean sheet to add panels onto.
    locs_path = _pm_cache_path(job_name, bp_name, pages)
    with open(locs_path, "w") as f:
        _json.dump({}, f)

    with open(_pm_session_path(job_name), "w") as f:
        _json.dump({"bp_name": bp_name, "folder": folder, "src_pdf": bp_path,
                    "scan_pdf": scan_path, "locs": locs_path, "pages": pages or []}, f)

    # Produce an (un-annotated) output so Download/Open still work before any edits.
    try:
        generate_panel_map_blueprint(scan_path, {}, _pm_output_path(job_name, bp_name),
                                     keep_only_panel_pages=False)
        _pm_build_full(job_name, bp_name, pages, bp_path)
        _pm_publish_to_viewer(job_name, bp_name)
    except Exception:
        pass

    return jsonify({"ok": True})


@app.route("/api/panel-map/status/<path:job_name>")
@login_required
def panel_map_status(job_name):
    with _pm_jobs_lock:
        job = dict(_pm_jobs.get(job_name, {}))
    bp_name = job.get("blueprint", "")
    job["has_output"] = bool(bp_name) and os.path.exists(_pm_output_path(job_name, bp_name))
    return jsonify(job)


@app.route("/api/panel-map/download/<path:job_name>")
@login_required
def panel_map_download(job_name):
    bp_name = (request.args.get("blueprint") or "").strip()
    if not bp_name:
        with _pm_jobs_lock:
            bp_name = _pm_jobs.get(job_name, {}).get("blueprint", "")
    out = _pm_output_path(job_name, bp_name) if bp_name else None
    if not out or not os.path.exists(out):
        return jsonify({"error": "No panel map yet"}), 404
    return send_file(out, mimetype="application/pdf", as_attachment=False,
                     download_name=f"{job_name} - Panel Map.pdf")


# ── Panel map editor (select / renumber / delete / add panels) ───────────────
@app.route("/panel-map/editor/<path:job_name>")
@login_required
def panel_map_editor(job_name):
    return render_template("panel_map_editor.html", job_name=job_name)


@app.route("/api/panel-map/editor-data/<path:job_name>")
@login_required
def panel_map_editor_data(job_name):
    sess = _pm_load_session(job_name)
    if not sess or not os.path.isfile(sess.get("scan_pdf", "")):
        return jsonify({"error": "Run the panel mapper first."}), 404
    locs = {}
    if os.path.isfile(sess.get("locs", "")):
        with open(sess["locs"]) as f:
            locs = _json.load(f)
    return jsonify({
        "job": job_name,
        "blueprint": sess.get("bp_name", ""),
        "pages": _pm_page_dims(sess["scan_pdf"]),
        "panel_locations": locs,
    })


@app.route("/api/panel-map/base/<path:job_name>")
@login_required
def panel_map_base(job_name):
    sess = _pm_load_session(job_name)
    if not sess or not os.path.isfile(sess.get("scan_pdf", "")):
        return jsonify({"error": "No base PDF"}), 404
    return send_file(sess["scan_pdf"], mimetype="application/pdf")


@app.route("/api/panel-map/ocr-region/<path:job_name>", methods=["POST"])
@login_required
def panel_map_ocr_region(job_name):
    """OCR a dragged rectangle on a page and return the panel numbers found."""
    from packing_list_engine import ocr_region
    sess = _pm_load_session(job_name)
    if not sess or not os.path.isfile(sess.get("scan_pdf", "")):
        return jsonify({"error": "Open the editor on a mapped document first."}), 404
    data = request.get_json(silent=True) or {}
    try:
        page = int(data.get("page"))
        bbox = [float(v) for v in data.get("bbox", [])][:4]
    except (TypeError, ValueError):
        return jsonify({"error": "Bad selection"}), 400
    if len(bbox) != 4:
        return jsonify({"error": "Bad selection"}), 400
    try:
        found = ocr_region(sess["scan_pdf"], page, bbox)
    except Exception as e:
        return jsonify({"error": f"Could not read that area: {e}"}), 500
    return jsonify({"ok": True, "panels": found})


@app.route("/api/panel-map/update/<path:job_name>", methods=["POST"])
@login_required
def panel_map_update(job_name):
    from packing_list_engine import generate_panel_map_blueprint
    sess = _pm_load_session(job_name)
    if not sess or not os.path.isfile(sess.get("scan_pdf", "")):
        return jsonify({"error": "Run the panel mapper first."}), 404
    locs = {}
    if os.path.isfile(sess.get("locs", "")):
        with open(sess["locs"]) as f:
            locs = _json.load(f)

    data = request.get_json(silent=True) or {}
    new_locs = data.get("locations")
    if not isinstance(new_locs, dict):
        return jsonify({"error": "Missing locations"}), 400

    clean = {}
    for panel, v in new_locs.items():
        p = str(panel).strip()
        if not p or not isinstance(v, dict):
            continue
        try:
            bbox = [float(x) for x in v.get("bbox", [])][:4]
            if len(bbox) != 4:
                continue
            entry = {"page": int(v.get("page")), "bbox": bbox}
            if v.get("label"):
                entry["label"] = str(v["label"])
            if v.get("rel"):
                entry["rel"] = str(v["rel"])    # 'sheet' = same panel on multiple sheets, 'dup' = duplicate number
            clean[p] = entry
        except (TypeError, ValueError):
            continue

    old_keys = set(locs.keys())
    new_keys = set(clean.keys())

    os.makedirs(_pm_dir(job_name), exist_ok=True)
    with open(sess["locs"], "w") as f:
        _json.dump(clean, f)

    # Regenerate the map PDF (marked-up pages only) + the full merged document.
    bp_name = sess.get("bp_name", "")
    result = generate_panel_map_blueprint(
        sess["scan_pdf"], clean, _pm_output_path(job_name, bp_name),
        keep_only_panel_pages=False)
    _pm_build_full(job_name, bp_name, sess.get("pages"), sess.get("src_pdf") or sess.get("scan_pdf"))
    _pm_publish_to_viewer(job_name, bp_name)

    # Log the human corrections to the shared cross-project record so the
    # packing-list tracker can benefit from how panels were dialed in.
    try:
        ts = datetime.utcnow().isoformat()
        with open(_GLOBAL_CORRECTIONS, "a") as gf:
            for p in (old_keys - new_keys):
                gf.write(_json.dumps({"type": "pm_remove", "panel": p, "job": job_name, "ts": ts}) + "\n")
            for p in (new_keys - old_keys):
                gf.write(_json.dumps({"type": "pm_add", "panel": p,
                                      "page": clean[p]["page"], "bbox": clean[p]["bbox"],
                                      "job": job_name, "ts": ts}) + "\n")
    except Exception:
        pass

    return jsonify({"ok": True, "panel_locations": clean,
                    "count": len(clean), "output_pages": result["output_pages"],
                    "added": len(new_keys - old_keys),
                    "removed": len(old_keys - new_keys),
                    "has_full": os.path.isfile(_pm_full_path(job_name, bp_name))})


@app.route("/api/panel-map/download-full/<path:job_name>")
@login_required
def panel_map_download_full(job_name):
    """Serve the full blueprint with the marked-up pages merged back in."""
    bp_name = (request.args.get("blueprint") or "").strip()
    if not bp_name:
        sess = _pm_load_session(job_name)
        bp_name = (sess or {}).get("bp_name", "")
    out = _pm_full_path(job_name, bp_name) if bp_name else None
    if not out or not os.path.exists(out):
        return jsonify({"error": "No full document yet"}), 404
    return send_file(out, mimetype="application/pdf", as_attachment=False,
                     download_name=f"{job_name} - Panel Map (Full).pdf")


@app.route("/packing-list-tracker")
@login_required
def packing_list_tracker():
    return render_template("packing_list_tracker.html")


@app.route("/api/packing-list/upload/<path:job_name>", methods=["POST"])
@login_required
def packing_list_upload(job_name):
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Must be a PDF"}), 400
    # Save to the Packing Lists folder so it appears in the file list
    pl_dir = safe_join(job_name, "Packing Lists")
    os.makedirs(pl_dir, exist_ok=True)
    os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
    label     = request.form.get("label", f.filename.replace(".pdf", ""))
    save_path = os.path.join(pl_dir, f.filename)
    f.save(save_path)
    with _pl_jobs_lock:
        _pl_jobs[job_name] = {"status": "processing", "progress": 0, "message": "Starting..."}
    _threading.Thread(target=_run_pl_job, args=(job_name, save_path, label), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/packing-list/process-file/<path:job_name>", methods=["POST"])
@login_required
def packing_list_process_file(job_name):
    data     = request.get_json(silent=True) or {}
    filename = data.get("filename", "").strip()
    if not filename or not filename.lower().endswith(".pdf"):
        return jsonify({"error": "Invalid filename"}), 400
    pl_path = safe_join(job_name, "Packing Lists", filename)
    if not os.path.isfile(pl_path):
        return jsonify({"error": f"File not found: {filename}"}), 404
    label = data.get("label") or filename.replace(".pdf", "")
    with _pl_jobs_lock:
        _pl_jobs[job_name] = {"status": "processing", "progress": 0, "message": "Starting..."}
    _threading.Thread(target=_run_pl_job, args=(job_name, pl_path, label), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/packing-list/status/<path:job_name>")
@login_required
def packing_list_status(job_name):
    with _pl_jobs_lock:
        job = dict(_pl_jobs.get(job_name, {}))
    if os.path.exists(_pl_state_path(job_name)):
        with open(_pl_state_path(job_name)) as f:
            state = _json.load(f)

        # Per-shipment stats. Color index comes from the persistent per-job map so
        # it matches the prints exactly and never changes.
        cmap = _pl_assign_colors(job_name, state)
        shipment_info = {}   # label -> {count, skids, color_index}
        for info in state.values():
            s = info.get("shipment", "Unknown")
            if s not in shipment_info:
                shipment_info[s] = {"count": 0, "skids": set(),
                                    "color_index": cmap.get(s, len(shipment_info))}
            shipment_info[s]["count"] += 1
            shipment_info[s]["skids"].add(info.get("skid", ""))

        # Merge in any shipments with 0 panels that were recorded in the registry
        reg_path = _pl_registry_path(job_name)
        if os.path.exists(reg_path):
            try:
                with open(reg_path) as f:
                    reg = _json.load(f)
                for label, rdata in reg.items():
                    if label not in shipment_info:
                        shipment_info[label] = {
                            "count": 0,
                            "skids": set(rdata.get("skids", [])),
                            "color_index": cmap.get(label, rdata.get("color_index", len(shipment_info))),
                        }
            except Exception:
                pass

        job["total_panels"] = len(state)
        _sk = lambda x: (int(x) if str(x).isdigit() else float('inf'), str(x))
        job["total_skids"]  = sorted(set(v["skid"] for v in state.values()), key=_sk)
        job["shipments"]    = [
            {"label": k, "count": v["count"],
             "skids": sorted(v["skids"], key=_sk),
             "color_index": v["color_index"]}
            for k, v in shipment_info.items()
        ]
        # Map filename -> color_index so the UI can highlight file items
        job["file_colors"]  = {
            s: v["color_index"] for s, v in shipment_info.items()
        }
    job["has_output"] = os.path.exists(_pl_output_path(job_name))
    return jsonify(job)


@app.route("/api/packing-list/download/<path:job_name>")
@login_required
def packing_list_download(job_name):
    output_path = _pl_output_path(job_name)
    if not os.path.exists(output_path):
        return jsonify({"error": "No tracked blueprint yet"}), 404
    return send_file(output_path, mimetype="application/pdf",
                     as_attachment=False,
                     download_name=f"{job_name}_delivery_tracked.pdf")


@app.route("/api/packing-list/publish/<path:job_name>", methods=["POST"])
@login_required
def packing_list_publish(job_name):
    output_path = _pl_output_path(job_name)
    if not os.path.exists(output_path):
        return jsonify({"error": "No tracked blueprint to publish"}), 404
    bp_dir = safe_join(job_name, "Blueprints")
    os.makedirs(bp_dir, exist_ok=True)
    used = {int(re.match(r'^(\d+)\s*-', fn).group(1))
            for fn in os.listdir(bp_dir) if re.match(r'^(\d+)\s*-', fn)}
    n = 1
    while n in used: n += 1
    dest = os.path.join(bp_dir, f"{n} - {job_name} Delivery Tracked.pdf")
    import shutil; shutil.copy2(output_path, dest)
    return jsonify({"ok": True, "filename": os.path.basename(dest), "number": n})


@app.route("/api/packing-list/reset/<path:job_name>", methods=["POST"])
@login_required
def packing_list_reset(job_name):
    for p in [_pl_state_path(job_name), _pl_output_path(job_name)]:
        if os.path.exists(p): os.unlink(p)
    with _pl_jobs_lock:
        _pl_jobs.pop(job_name, None)
    return jsonify({"ok": True})


@app.route("/api/packing-list/unlink/<path:job_name>", methods=["POST"])
@login_required
def packing_list_unlink(job_name):
    """Remove one or more shipments' tracking data from the job — panel highlights,
    delivery state entries, registry entries, and color assignments — without
    deleting the source PDF files.  Regenerates the tracked blueprint afterward."""
    data      = request.get_json(silent=True) or {}
    filenames = data.get("filenames", [])
    if not filenames:
        return jsonify({"error": "No filenames provided"}), 400

    # Derive shipment labels from filenames (label == filename minus .pdf extension)
    labels_to_remove = set()
    for fname in filenames:
        label = re.sub(r"\.pdf$", "", fname, flags=re.IGNORECASE).strip()
        if label:
            labels_to_remove.add(label)

    if not labels_to_remove:
        return jsonify({"error": "No valid shipment labels derived"}), 400

    # ── Strip from delivery_state ─────────────────────────────────────────
    delivery_state = {}
    if os.path.exists(_pl_state_path(job_name)):
        with open(_pl_state_path(job_name)) as f:
            delivery_state = _json.load(f)
    delivery_state = {p: v for p, v in delivery_state.items()
                      if v.get("shipment") not in labels_to_remove}
    with open(_pl_state_path(job_name), "w") as f:
        _json.dump(delivery_state, f)

    # ── Strip from shipment registry ──────────────────────────────────────
    reg_path = _pl_registry_path(job_name)
    if os.path.exists(reg_path):
        with open(reg_path) as f:
            reg = _json.load(f)
        reg = {k: v for k, v in reg.items() if k not in labels_to_remove}
        with open(reg_path, "w") as f:
            _json.dump(reg, f)

    # ── Strip from ship_colors ────────────────────────────────────────────
    colors_path = _pl_colors_path(job_name)
    if os.path.exists(colors_path):
        with open(colors_path) as f:
            cmap = _json.load(f)
        cmap = {k: v for k, v in cmap.items() if k not in labels_to_remove}
        with open(colors_path, "w") as f:
            _json.dump(cmap, f)
    else:
        cmap = {}

    # ── Rebuild ship_colors for remaining shipments ───────────────────────
    ship_colors = _pl_assign_colors(job_name, delivery_state)

    # ── Regenerate tracked blueprint ─────────────────────────────────────
    if delivery_state:
        try:
            from packing_list_engine import generate_tracked_blueprint, generate_tracked_blueprint_panel_map
            pm_sess      = _pm_load_session(job_name)
            pm_locs_path = pm_sess.get("locs", "") if pm_sess else ""
            pm_scan_pdf  = pm_sess.get("scan_pdf", "") if pm_sess else ""
            use_panel_map = bool(pm_sess and pm_locs_path and os.path.isfile(pm_locs_path)
                                 and pm_scan_pdf and os.path.isfile(pm_scan_pdf))
            if use_panel_map:
                with open(pm_locs_path) as f:
                    panel_locations = _json.load(f)
                generate_tracked_blueprint_panel_map(
                    pm_scan_pdf, panel_locations, delivery_state, _pl_output_path(job_name),
                    shipment_colors=ship_colors)
            else:
                panel_locations = {}
                if os.path.exists(_pl_cache_path(job_name)):
                    with open(_pl_cache_path(job_name)) as f:
                        panel_locations = _json.load(f)
                blueprint_path = _find_blueprint(job_name)
                if blueprint_path and panel_locations:
                    generate_tracked_blueprint(
                        blueprint_path, delivery_state, panel_locations, _pl_output_path(job_name),
                        shipment_colors=ship_colors)
        except Exception as e:
            logging.warning(f"packing_list_unlink: blueprint regen failed: {e}")
    else:
        # No shipments left — remove the output blueprint entirely
        if os.path.exists(_pl_output_path(job_name)):
            os.unlink(_pl_output_path(job_name))

    with _pl_jobs_lock:
        _pl_jobs.pop(job_name, None)

    return jsonify({"ok": True, "unlinked": list(labels_to_remove)})


@app.route("/api/packing-list/manual-panels/<path:job_name>", methods=["POST"])
@login_required
def packing_list_manual_panels(job_name):
    """Add panels to a shipment manually (for handwritten or unreadable packing lists)."""
    data = request.get_json(force=True) or {}
    label = data.get("label", "").strip()
    rows  = data.get("rows", [])
    if not label or not rows:
        return jsonify({"error": "label and rows required"}), 400

    os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
    delivery_state = {}
    if os.path.exists(_pl_state_path(job_name)):
        with open(_pl_state_path(job_name)) as f:
            delivery_state = _json.load(f)

    _panel_re = re.compile(r"^\d{1,3}[A-Za-z]?$")
    added = 0
    parsed_skids = []

    for row in rows:
        skid   = str(row.get("skid", "")).strip() or "?"
        panels_text = str(row.get("panels", ""))
        parsed_skids.append(skid)
        tokens = re.split(r"[,\s]+", panels_text)
        for tok in tokens:
            tok = tok.strip()
            if not tok:
                continue
            # Normalize: strip leading zeros → "05" → "5"
            try:
                norm = str(int(re.match(r"(\d+)", tok).group(1)))
                suffix = re.sub(r"^\d+", "", tok)
                tok = norm + suffix.upper()
            except Exception:
                continue
            if not _panel_re.match(tok):
                continue
            try:
                num = int(re.match(r"(\d+)", tok).group(1))
                if not (1 <= num <= 700):
                    continue
            except Exception:
                continue
            if tok not in delivery_state:
                added += 1
            elif delivery_state[tok].get("shipment") != label:
                added += 1  # count reassignments too so UI shows something meaningful
            delivery_state[tok] = {"skid": skid, "shipment": label, "order_num": ""}

    with open(_pl_state_path(job_name), "w") as f:
        _json.dump(delivery_state, f)

    ship_colors = _pl_assign_colors(job_name, delivery_state, extra_shipments=[label])
    _pl_record_shipment(job_name, label, parsed_skids, ship_colors)

    # Regenerate blueprint
    try:
        from packing_list_engine import generate_tracked_blueprint, generate_tracked_blueprint_panel_map
        pm_sess      = _pm_load_session(job_name)
        pm_locs_path = pm_sess.get("locs", "") if pm_sess else ""
        pm_scan_pdf  = pm_sess.get("scan_pdf", "") if pm_sess else ""
        use_panel_map = bool(pm_sess and pm_locs_path and os.path.isfile(pm_locs_path)
                             and pm_scan_pdf and os.path.isfile(pm_scan_pdf))
        if use_panel_map:
            with open(pm_locs_path) as f:
                panel_locations = _json.load(f)
            try:
                with open(_pl_cache_path(job_name), "w") as f:
                    _json.dump(panel_locations, f)
            except Exception:
                pass
            generate_tracked_blueprint_panel_map(
                pm_scan_pdf, panel_locations, delivery_state, _pl_output_path(job_name),
                shipment_colors=ship_colors)
        else:
            panel_locations = {}
            if os.path.exists(_pl_cache_path(job_name)):
                with open(_pl_cache_path(job_name)) as f:
                    panel_locations = _json.load(f)
            blueprint_path = _find_blueprint(job_name)
            if blueprint_path and panel_locations:
                generate_tracked_blueprint(
                    blueprint_path, delivery_state, panel_locations, _pl_output_path(job_name),
                    shipment_colors=ship_colors)
    except Exception:
        pass  # state saved; blueprint regen failed — user can reprocess

    return jsonify({"ok": True, "added": added})


# ── Packing List Editor (interactive cross-reference) ─────────────────────────

_sk_key = lambda x: (int(x) if str(x).isdigit() else float('inf'), str(x))

def _pl_stats(state, color_map=None):
    """Per-shipment stats. Color index comes from the persistent per-job color_map
    when provided (matches the prints); otherwise first-seen order."""
    color_map = color_map or {}
    shipment_info = {}
    for info in state.values():
        s = info.get("shipment", "Unknown")
        if s not in shipment_info:
            shipment_info[s] = {"count": 0, "skids": set(),
                                "color_index": color_map.get(s, len(shipment_info))}
        shipment_info[s]["count"] += 1
        shipment_info[s]["skids"].add(str(info.get("skid", "")))
    shipments = [
        {"label": k, "count": v["count"],
         "skids": sorted(v["skids"], key=_sk_key),
         "color_index": v["color_index"]}
        for k, v in shipment_info.items()
    ]
    return {
        "total_panels": len(state),
        "total_skids": sorted({str(v.get("skid", "")) for v in state.values()}, key=_sk_key),
        "shipments": shipments,
        "file_colors": {s: v["color_index"] for s, v in shipment_info.items()},
    }

def _pl_load_state(job_name):
    if os.path.exists(_pl_state_path(job_name)):
        with open(_pl_state_path(job_name)) as f:
            return _json.load(f)
    return {}

def _pl_load_locations(job_name):
    if os.path.exists(_pl_cache_path(job_name)):
        with open(_pl_cache_path(job_name)) as f:
            return _json.load(f)
    return {}

def _pl_load_cells(job_name):
    if os.path.exists(_pl_cells_path(job_name)):
        try:
            with open(_pl_cells_path(job_name)) as f:
                return _json.load(f)
        except Exception:
            pass
    return {}

# ── Manual corrections store ─────────────────────────────────────────────────
# Captures every hand-fix made in the editor (delete a false positive, renumber a
# panel, add a missing one). Stored per job so the fixes survive re-processing,
# and appended to a single cross-project log so future panel-finding work can
# learn from how humans corrected the automatic results.
_GLOBAL_CORRECTIONS = os.path.join(JOBS_DIR, ".panel_corrections.jsonl")

def _pl_corrections_path(job_name):
    return os.path.join(_pl_tracking_dir(job_name), "corrections.json")

def _pl_load_corrections(job_name):
    c = {}
    p = _pl_corrections_path(job_name)
    if os.path.exists(p):
        try:
            with open(p) as f:
                c = _json.load(f)
        except Exception:
            c = {}
    c.setdefault("deletions", [])
    c.setdefault("renames", {})
    c.setdefault("additions", {})
    return c

def _pl_save_corrections(job_name, c):
    try:
        os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
        with open(_pl_corrections_path(job_name), "w") as f:
            _json.dump(c, f)
    except Exception:
        pass

def _log_global_correction(job_name, event):
    """Append one correction event to the cross-project corrections log."""
    try:
        ev = dict(event)
        ev["job"] = job_name
        ev["ts"] = datetime.utcnow().isoformat() + "Z"
        with open(_GLOBAL_CORRECTIONS, "a") as f:
            f.write(_json.dumps(ev) + "\n")
    except Exception:
        pass

def _apply_corrections(delivery_state, panel_locations, corr):
    """Re-apply a job's saved manual corrections after an automatic scan, so the
    user's fixes are not undone by re-processing."""
    for old, new in (corr.get("renames") or {}).items():
        if old in panel_locations:
            panel_locations[new] = panel_locations.pop(old)
        if old in delivery_state:
            delivery_state[new] = delivery_state.pop(old)
    for p in (corr.get("deletions") or []):
        panel_locations.pop(p, None)
        delivery_state.pop(p, None)
    for p, info in (corr.get("additions") or {}).items():
        if info.get("page") is not None and info.get("bbox"):
            panel_locations[p] = {"page": int(info["page"]), "bbox": info["bbox"]}
        if p not in delivery_state and info.get("shipment"):
            delivery_state[p] = {"skid": info.get("skid", ""), "shipment": info["shipment"]}

def _pl_blueprint_dims(job_name):
    """Per-page [{width,height}] of the job blueprint, in PDF points."""
    bp = _find_blueprint(job_name)
    if not bp:
        return [], None
    import fitz as _fitz
    doc = _fitz.open(bp)
    dims = [{"width": p.rect.width, "height": p.rect.height} for p in doc]
    doc.close()
    return dims, bp


@app.route("/packing-list/editor/<path:job_name>")
@login_required
def packing_list_editor(job_name):
    return render_template("packing_list_editor.html", job_name=job_name)


@app.route("/api/packing-list/editor-data/<path:job_name>")
@login_required
def packing_list_editor_data(job_name):
    """Everything the editor needs to render the blueprint with panel overlays."""
    state = _pl_load_state(job_name)
    locations = _pl_load_locations(job_name)

    # Merge Panel Mapper panel locations as a fallback so the editor can cross-link
    # panels that the packing-list DXF scan didn't locate (the red boxes visible on
    # the blueprint come from the Panel Mapper scan; without this those panels show as
    # "unlocated" in the editor even though their position is known).
    pm_sess = _pm_load_session(job_name)
    if pm_sess:
        pm_locs_path = pm_sess.get("locs", "")
        if pm_locs_path and os.path.isfile(pm_locs_path):
            try:
                with open(pm_locs_path) as f:
                    pm_locs = _json.load(f)
                for panel, loc in pm_locs.items():
                    if panel not in locations:
                        locations[panel] = loc
            except Exception:
                pass

    dims, bp = _pl_blueprint_dims(job_name)
    if not bp:
        return jsonify({"error": "No blueprint PDF found for this job."}), 404

    # Available packing list files (for the right pane selector)
    pl_files = []
    pl_dir = safe_join(job_name, "Packing Lists")
    if os.path.isdir(pl_dir):
        pl_files = sorted(f for f in os.listdir(pl_dir) if f.lower().endswith(".pdf"))

    # If a Panel Mapper session exists, expose the panels_only URL so the editor
    # can show that file instead of the full tracked blueprint — page indices match.
    panels_only_url = None
    pm_sess = _pm_load_session(job_name)
    if pm_sess:
        bp_name = pm_sess.get("bp_name", "")
        if bp_name:
            base = re.sub(r'\s*-\s*Panel Mapper$', '', bp_name, flags=re.IGNORECASE)
            base = re.sub(r'\.pdf$', '', base, flags=re.IGNORECASE)
            po_name = f"{base} - panels_only.pdf"
            po_path = safe_join(job_name, "Panel Mapper", po_name)
            if os.path.isfile(po_path):
                from urllib.parse import quote as _uq
                import time as _time
                panels_only_url = f"/files/{_uq(job_name)}/Panel Mapper/{_uq(po_name)}?t={int(_time.time())}"

    stats = _pl_stats(state, _pl_assign_colors(job_name, state))
    return jsonify({
        "job": job_name,
        "pages": dims,
        "panel_locations": locations,   # {panel: {page, bbox}}  ALL panels on blueprint
        "delivery_state": state,        # {panel: {skid, shipment}} delivered only
        "table_cells": _pl_load_cells(job_name),  # {panel: {page, bbox}} DELIVERED table row
        "packing_lists": pl_files,
        "panels_only_url": panels_only_url,  # preferred view: Panel Mapper panels_only.pdf
        **stats,
        "has_output": os.path.exists(_pl_output_path(job_name)),
    })


@app.route("/api/packing-list/blueprint/<path:job_name>")
@login_required
def packing_list_blueprint(job_name):
    """Serve the base (un-annotated) blueprint PDF — overlays are drawn client-side."""
    bp = _find_blueprint(job_name)
    if not bp or not os.path.exists(bp):
        return jsonify({"error": "No blueprint PDF found"}), 404
    return send_file(bp, mimetype="application/pdf")


def _pl_safe_list_file(job_name, filename):
    if not filename or not filename.lower().endswith(".pdf"):
        return None
    p = safe_join(job_name, "Packing Lists", filename)
    return p if os.path.isfile(p) else None


@app.route("/api/packing-list/pl-file/<path:job_name>")
@login_required
def packing_list_pl_file(job_name):
    """Serve a single packing list PDF for the editor's right pane."""
    p = _pl_safe_list_file(job_name, request.args.get("file", ""))
    if not p:
        return jsonify({"error": "File not found"}), 404
    return send_file(p, mimetype="application/pdf")


@app.route("/api/packing-list/pl-positions/<path:job_name>")
@login_required
def packing_list_pl_positions(job_name):
    """OCR a packing list for panel-number positions (cached) for cross-highlighting."""
    filename = request.args.get("file", "")
    p = _pl_safe_list_file(job_name, filename)
    if not p:
        return jsonify({"error": "File not found"}), 404
    try:
        from packing_list_engine import scan_packing_list_positions
        safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", filename)
        # v2 = PANEL #-column-aware scan (invalidates older full-page caches)
        cache = os.path.join(_pl_tracking_dir(job_name), f"pl_pos_v4_{safe_name}.json")
        os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
        data = scan_packing_list_positions(p, cache_path=cache)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/packing-list/update-panels/<path:job_name>", methods=["POST"])
@login_required
def packing_list_update_panels(job_name):
    """Add/remove/renumber delivered panels, then regenerate the tracked blueprint.

    Body: {
      "add":     [{"panel","skid","shipment","page"?,"bbox"?}, ...],
      "remove":  ["panel", ...],
      "renames": [{"from","to","skid"?,"shipment"?}, ...]
    }
    All edits are also recorded in the per-job + global corrections store.
    """
    data = request.get_json(silent=True) or {}
    to_add = data.get("add", []) or []
    to_remove = data.get("remove", []) or []
    renames = data.get("renames", []) or []

    state = _pl_load_state(job_name)
    locations = _pl_load_locations(job_name)
    corr = _pl_load_corrections(job_name)

    # ── Renames (renumber a panel; the table will then show the correct number) ──
    for r in renames:
        old = str(r.get("from", "")).strip()
        new = str(r.get("to", "")).strip()
        if not old or not new or old == new:
            continue
        if old in state:
            state[new] = state.pop(old)
        if old in locations:
            locations[new] = locations.pop(old)
        corr["renames"][old] = new
        if old in corr["additions"]:
            corr["additions"][new] = corr["additions"].pop(old)
        if old in corr["deletions"]:
            corr["deletions"].remove(old)
        _log_global_correction(job_name, {
            "type": "rename", "from": old, "to": new,
            "page": (locations.get(new) or {}).get("page"),
            "bbox": (locations.get(new) or {}).get("bbox")})

    removed = 0
    for panel in to_remove:
        if panel in state:
            del state[panel]
            removed += 1
        loc = locations.pop(panel, None)   # also drop the highlight location (false positives)
        if panel not in corr["deletions"]:
            corr["deletions"].append(panel)
        corr["additions"].pop(panel, None)
        _log_global_correction(job_name, {
            "type": "delete", "panel": panel,
            "page": (loc or {}).get("page"), "bbox": (loc or {}).get("bbox")})

    added = 0
    for item in to_add:
        panel = str(item.get("panel", "")).strip()
        if not panel:
            continue
        skid = str(item.get("skid", "")).strip()
        shipment = str(item.get("shipment", "")).strip() or "Manual"
        state[panel] = {"skid": skid, "shipment": shipment}
        page = item.get("page")
        bbox = None
        # If the panel wasn't located by OCR, store the box the user drew so the
        # regenerated PDF highlights it too.
        if item.get("bbox") and page is not None:
            try:
                bbox = [float(v) for v in item["bbox"]]
                locations[panel] = {"page": int(page), "bbox": bbox}
            except Exception:
                bbox = None
        # Record as a manual addition so a future re-scan keeps it.
        corr["additions"][panel] = {"panel": panel, "skid": skid, "shipment": shipment,
                                    "page": (int(page) if page is not None else None), "bbox": bbox}
        if panel in corr["deletions"]:
            corr["deletions"].remove(panel)
        _log_global_correction(job_name, {
            "type": "add", "panel": panel, "skid": skid, "shipment": shipment,
            "page": (int(page) if page is not None else None), "bbox": bbox})
        added += 1

    _pl_save_corrections(job_name, corr)

    os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
    with open(_pl_state_path(job_name), "w") as f:
        _json.dump(state, f)
    with open(_pl_cache_path(job_name), "w") as f:
        _json.dump(locations, f)

    # Regenerate the annotated PDF so Download / Publish reflect the edits.
    # Use the same path as _run_pl_job: Panel Mapper scan_pdf when available,
    # otherwise fall back to the full original blueprint.
    regen_ok, regen_err = True, None
    table_cells = {}
    try:
        from packing_list_engine import generate_tracked_blueprint, generate_tracked_blueprint_panel_map
        ship_colors = _pl_assign_colors(job_name, state)
        pm_sess = _pm_load_session(job_name)
        pm_locs_path = pm_sess.get("locs", "") if pm_sess else ""
        pm_scan_pdf  = pm_sess.get("scan_pdf", "") if pm_sess else ""
        use_panel_map = bool(pm_sess and pm_locs_path and os.path.isfile(pm_locs_path)
                             and pm_scan_pdf and os.path.isfile(pm_scan_pdf))
        if use_panel_map:
            with open(pm_locs_path) as f:
                panel_locations_pm = _json.load(f)
            merged_locations = dict(panel_locations_pm)
            merged_locations.update(locations)
            generate_tracked_blueprint_panel_map(
                pm_scan_pdf, merged_locations, state, _pl_output_path(job_name),
                shipment_colors=ship_colors)
        else:
            blueprint_path = _find_blueprint(job_name)
            if blueprint_path and locations:
                generate_tracked_blueprint(
                    blueprint_path, state, locations, _pl_output_path(job_name),
                    shipment_colors=ship_colors)
    except Exception as e:
        regen_ok = False
        regen_err = str(e)

    stats = _pl_stats(state, _pl_assign_colors(job_name, state))
    return jsonify({
        "ok": True,
        "added": added,
        "removed": removed,
        "regenerated": regen_ok,
        "regen_error": regen_err,
        "delivery_state": state,
        "panel_locations": locations,
        "table_cells": table_cells,
        "shipments": stats.get("shipments", []),
        "file_colors": stats.get("file_colors", {}),
    })


# ════════════════════════════════════════════════════════════════════
# Panel Tracking — combined tool (NEW pages only; reuses existing APIs
# and the same stored job data so shipment colors stay constant).
# ════════════════════════════════════════════════════════════════════

@app.route("/panel-tracking")
@login_required
def panel_tracking_home():
    return render_template("panel_tracking.html", pt_job=None, pt_tab="overview")


@app.route("/panel-tracking/<path:job_name>/map")
@login_required
def panel_tracking_map(job_name):
    return render_template("pt_map.html", pt_job=job_name, pt_tab="map")


@app.route("/panel-tracking/<path:job_name>/map/editor")
@login_required
def panel_tracking_map_editor(job_name):
    return render_template("pt_map_editor.html", job_name=job_name, pt_job=job_name)


@app.route("/panel-tracking/<path:job_name>/deliveries")
@login_required
def panel_tracking_deliveries(job_name):
    return render_template("pt_deliveries.html", pt_job=job_name, pt_tab="deliveries")


@app.route("/panel-tracking/<path:job_name>/review")
@login_required
def panel_tracking_review(job_name):
    return render_template("pt_review.html", job_name=job_name, pt_job=job_name)


@app.route("/panel-tracking/<path:job_name>/documents")
@login_required
def panel_tracking_documents(job_name):
    return render_template("pt_documents.html", pt_job=job_name, pt_tab="documents")


@app.route("/panel-tracking/<path:job_name>/fab")
@login_required
def panel_tracking_fab(job_name):
    return render_template("pt_fab.html", pt_job=job_name, pt_tab="fab")


@app.route("/panel-tracking/<path:job_name>")
@login_required
def panel_tracking_overview(job_name):
    return render_template("panel_tracking.html", pt_job=job_name, pt_tab="overview")


# ════════════════════════════════════════════════════════════════════
# Public Links — one permanent public link per document slot, per job.
# Publish swaps the document behind a link (link never changes).
# "New Link" rotates the token — the old link stops working instantly.
# Store: <JOBS_DIR>/.public_links.json = {job: {slot: {token, file, published}}}
# ════════════════════════════════════════════════════════════════════

PUBLIC_LINKS_FILE = os.path.join(JOBS_DIR, ".public_links.json")
_PUB_SLOTS = ("prints", "panels_only", "packing_lists", "fab_sheets",
              "spreadsheet", "todo_board", "job_page")
_PUB_SLOT_NAMES = {
    "prints":        "Marked-Up Prints",
    "panels_only":   "Panels-Only Prints",
    "packing_lists": "All Packing Lists",
    "fab_sheets":    "All Fab Sheets",
    "spreadsheet":   "Job Spreadsheet",
    "todo_board":    "To-Do Board",
    "job_page":      "Job Page",
}
# Slots whose document is picked by hand at Publish time
_PUB_PICKED_SLOTS = ("prints", "panels_only")
# Slots that always serve the current folder contents (auto-update)
_PUB_AUTO_SLOTS = ("packing_lists", "fab_sheets", "spreadsheet")


def _pub_load():
    if os.path.exists(PUBLIC_LINKS_FILE):
        try:
            with open(PUBLIC_LINKS_FILE) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _pub_save(data):
    with open(PUBLIC_LINKS_FILE, "w") as f:
        json.dump(data, f, indent=2)


def _pub_find(token):
    """Resolve a public token -> (job, slot, info). Token must match EXACTLY
    the currently stored one — rotated-away tokens resolve to nothing."""
    if not token:
        return None, None, None
    data = _pub_load()
    for job, slots in data.items():
        for slot, info in slots.items():
            if isinstance(info, dict) and info.get("token") == token:
                return job, slot, info
    return None, None, None


def _pub_merged_pdf(job_name, folder, cache_name):
    """Combine every PDF in <job>/<folder> into one document (with a bookmark
    per source file). Cached in <job>/Public Links/ and rebuilt automatically
    whenever the source folder's contents change."""
    import fitz
    src_dir = safe_join(job_name, folder)
    if not os.path.isdir(src_dir):
        return None
    files = sorted(f for f in os.listdir(src_dir) if f.lower().endswith(".pdf"))
    if not files:
        return None

    out_dir = safe_join(job_name, "Public Links")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, cache_name)
    meta_path = out_path + ".meta.json"

    sig = [[f, os.path.getmtime(os.path.join(src_dir, f))] for f in files]
    if os.path.exists(out_path) and os.path.exists(meta_path):
        try:
            with open(meta_path) as f:
                if json.load(f) == sig:
                    return out_path          # cache is current
        except Exception:
            pass

    doc = fitz.open()
    toc = []
    for f in files:
        try:
            src = fitz.open(os.path.join(src_dir, f))
            toc.append([1, os.path.splitext(f)[0], doc.page_count + 1])
            doc.insert_pdf(src)
            src.close()
        except Exception:
            continue                          # skip unreadable PDFs
    if not doc.page_count:
        doc.close()
        return None
    try:
        doc.set_toc(toc)
    except Exception:
        pass
    doc.save(out_path)
    doc.close()
    with open(meta_path, "w") as f:
        json.dump(sig, f)
    return out_path


def _pub_spreadsheet_path(job_name):
    """Current xlsx for the job — prefers <job>.xlsx, else first xlsx found."""
    ss_dir = safe_join(job_name, "Spreadsheets")
    if not os.path.isdir(ss_dir):
        return None
    preferred = os.path.join(ss_dir, f"{job_name}.xlsx")
    if os.path.isfile(preferred):
        return preferred
    xs = sorted(f for f in os.listdir(ss_dir)
                if f.lower().endswith((".xlsx", ".xls")))
    return os.path.join(ss_dir, xs[0]) if xs else None


def _pub_sheets_url(job_name):
    """The job's linked Google Sheets URL (same file the Documents tab uses)."""
    p = safe_join(job_name, "Spreadsheets", f"{job_name}_sheets_url.json")
    if os.path.isfile(p):
        try:
            with open(p) as f:
                return (json.load(f) or {}).get("url", "")
        except Exception:
            pass
    return ""


def _pub_current_file(job_name, slot, info):
    """Absolute path of the document currently behind a slot, or None."""
    if slot in _PUB_PICKED_SLOTS:
        rel = (info or {}).get("file")
        if not rel:
            return None
        path = safe_join(job_name, rel)
        return path if os.path.isfile(path) else None
    if slot == "packing_lists":
        return _pub_merged_pdf(job_name, "Packing Lists", "All Packing Lists.pdf")
    if slot == "fab_sheets":
        return _pub_merged_pdf(job_name, "Fab Sheets", "All Fab Sheets.pdf")
    if slot == "spreadsheet":
        return _pub_spreadsheet_path(job_name)
    return None


# ── Admin: the tool page ─────────────────────────────────────────────

@app.route("/public-links")
@login_required
def public_links_tool():
    return render_template("public_links.html")


@app.route("/api/public-links/<path:job_name>")
@login_required
def public_links_state(job_name):
    job_path = safe_join(job_name)
    if not os.path.isdir(job_path):
        return jsonify({"error": "Job not found"}), 404
    job_links = _pub_load().get(job_name, {})

    # Candidate PDFs for the hand-picked print slots
    candidates = []
    for folder in ("Blueprints", "Panel Mapper"):
        fdir = os.path.join(job_path, folder)
        if os.path.isdir(fdir):
            for f in sorted(os.listdir(fdir)):
                if f.lower().endswith(".pdf"):
                    candidates.append({"folder": folder, "name": f})

    def _count_pdfs(folder):
        fdir = os.path.join(job_path, folder)
        if not os.path.isdir(fdir):
            return 0
        return len([f for f in os.listdir(fdir) if f.lower().endswith(".pdf")])

    ss_path = _pub_spreadsheet_path(job_name)
    slots = {}
    for slot in _PUB_SLOTS:
        info = job_links.get(slot) or {}
        live = bool(info.get("token"))
        if slot == "todo_board" and not live and \
           not _load_webhooks().get(job_name, {}).get("todo"):
            continue  # only jobs with a to-do channel get the board slot
        slots[slot] = {
            "name": _PUB_SLOT_NAMES[slot],
            "live": live,
            "url": (request.url_root.rstrip("/") + "/pl/" + info["token"]) if live else None,
            "file": info.get("file"),
            "published": info.get("published"),
        }
    return jsonify({
        "slots": slots,
        "candidates": candidates,
        "packing_count": _count_pdfs("Packing Lists"),
        "fab_count": _count_pdfs("Fab Sheets"),
        "spreadsheet_file": os.path.basename(ss_path) if ss_path else None,
        "sheets_url": _pub_sheets_url(job_name),
    })


@app.route("/api/public-links/<path:job_name>/publish", methods=["POST"])
@login_required
def public_links_publish(job_name):
    """Designate (or swap) the document behind a slot. Creates the link the
    first time; afterwards the SAME link simply starts serving the new doc."""
    job_path = safe_join(job_name)
    if not os.path.isdir(job_path):
        return jsonify({"error": "Job not found"}), 404
    body = request.get_json(silent=True) or {}
    slot = body.get("slot", "")
    if slot not in _PUB_SLOTS:
        return jsonify({"error": "Unknown slot"}), 400

    data = _pub_load()
    info = data.setdefault(job_name, {}).setdefault(slot, {})

    if slot in _PUB_PICKED_SLOTS:
        folder = body.get("folder", "")
        fname = body.get("file", "")
        if folder not in ("Blueprints", "Panel Mapper") or not fname:
            return jsonify({"error": "Pick a PDF to publish"}), 400
        path = safe_join(job_name, folder, fname)
        if not os.path.isfile(path):
            return jsonify({"error": "File not found"}), 404
        info["file"] = f"{folder}/{fname}"
    elif slot == "todo_board":
        # No document — the /pl link forwards to THIS job's to-do board
        if not _load_webhooks().get(job_name, {}).get("todo"):
            return jsonify({"error": f"No to-do channel is set up for "
                            f"“{job_name}” yet — create its Discord server "
                            f"and paste its to-do webhook first."}), 400
        try:
            _todo_token_for(job_name, create=True)
        except Exception:
            return jsonify({"error": "Could not set up the To-Do viewer "
                            "on the server."}), 500
    else:
        # Auto slots: just verify there is something to serve right now
        if not _pub_current_file(job_name, slot, info) and slot != "job_page":
            return jsonify({"error": "Nothing to publish yet — the source "
                            "folder is empty."}), 400

    if not info.get("token"):
        info["token"] = secrets.token_urlsafe(16)
    now = datetime.now()
    info["published"] = f"{now.month}/{now.day}/{now.year % 100} {now:%H:%M}"
    _pub_save(data)
    return jsonify({"ok": True,
                    "url": request.url_root.rstrip("/") + "/pl/" + info["token"]})


@app.route("/api/public-links/<path:job_name>/rotate", methods=["POST"])
@login_required
def public_links_rotate(job_name):
    """Generate a NEW link for a slot. The old link stops working instantly."""
    body = request.get_json(silent=True) or {}
    slot = body.get("slot", "")
    if slot not in _PUB_SLOTS:
        return jsonify({"error": "Unknown slot"}), 400
    data = _pub_load()
    info = data.setdefault(job_name, {}).setdefault(slot, {})
    info["token"] = secrets.token_urlsafe(16)
    now = datetime.now()
    info["published"] = f"{now.month}/{now.day}/{now.year % 100} {now:%H:%M}"
    _pub_save(data)
    return jsonify({"ok": True,
                    "url": request.url_root.rstrip("/") + "/pl/" + info["token"]})


@app.route("/api/public-links/<path:job_name>/disable", methods=["POST"])
@login_required
def public_links_disable(job_name):
    """Kill a slot's link without creating a new one. The chosen document is
    remembered, so re-publishing restores the same doc under a fresh link."""
    body = request.get_json(silent=True) or {}
    slot = body.get("slot", "")
    if slot not in _PUB_SLOTS:
        return jsonify({"error": "Unknown slot"}), 400
    data = _pub_load()
    info = data.get(job_name, {}).get(slot)
    if info:
        info.pop("token", None)
        _pub_save(data)
    return jsonify({"ok": True})


@app.route("/api/qr")
@login_required
def qr_png():
    """PNG QR code for a public link (Public Links page QR buttons)."""
    data = request.args.get("data", "")
    if not data or len(data) > 500:
        abort(400)
    try:
        import qrcode
    except ImportError:
        return jsonify({"error": "Server missing the 'qrcode' library — run the full deploy (deploy.bat) once."}), 500
    img = qrcode.make(data, box_size=10, border=2)
    buf = io.BytesIO()
    img.save(buf, "PNG")
    buf.seek(0)
    fname = (request.args.get("name") or "qr") + ".png"
    return send_file(buf, mimetype="image/png",
                     as_attachment=(request.args.get("dl") == "1"),
                     download_name=fname)


# ── Public: no login required — the token IS the access ─────────────

@app.route("/pl/<token>")
def public_link_view(token):
    job, slot, info = _pub_find(token)
    if not job:
        abort(404)

    if slot == "todo_board":
        view_tok = _todo_token_for(job)
        if not view_tok:
            abort(404)
        return redirect("/todo/" + view_tok)

    if slot == "job_page":
        job_links = _pub_load().get(job, {})
        docs = []
        for s in _PUB_SLOTS:
            if s == "job_page":
                continue
            i = job_links.get(s) or {}
            has_doc = (s == "todo_board") or _pub_current_file(job, s, i)
            if i.get("token") and has_doc:
                docs.append({"slot": s, "name": _PUB_SLOT_NAMES[s],
                             "url": "/pl/" + i["token"]})
        return render_template("public_job_page.html", job=job, docs=docs)

    path = _pub_current_file(job, slot, info)
    if not path:
        abort(404)

    if slot == "spreadsheet":
        # If the All-Packing-Lists link is live, packing-list cells link to it.
        # Panel-number cells: logged-in users go to Review & Fix; everyone else
        # goes to the public Panels-Only prints (zoomed to the panel).
        job_links = _pub_load().get(job, {})
        pl_info = job_links.get("packing_lists") or {}
        pl_url = ("/pl/" + pl_info["token"]) if pl_info.get("token") else ""
        po_info = job_links.get("panels_only") or {}
        po_url = ("/pl/" + po_info["token"]) if po_info.get("token") else ""
        return render_template("public_sheet_view.html", job=job, token=token,
                               fname=os.path.basename(path),
                               sheets_url=_pub_sheets_url(job),
                               pl_public_url=pl_url,
                               panels_only_url=po_url,
                               logged_in=bool(session.get("user")))
    return render_template("public_doc_view.html", job=job, token=token,
                           title=_PUB_SLOT_NAMES.get(slot, "Document"),
                           fname=os.path.basename(path))


@app.route("/pl/<token>/file")
def public_link_file(token):
    job, slot, info = _pub_find(token)
    if not job or slot == "job_page":
        abort(404)
    path = _pub_current_file(job, slot, info)
    if not path:
        abort(404)
    return send_file(path,
                     as_attachment=(request.args.get("dl") == "1"),
                     download_name=os.path.basename(path))


@app.route("/api/pt/fix-missed/<path:job_name>", methods=["POST"])
@login_required
def pt_fix_missed(job_name):
    """Review-tab 'Fix Missed': deterministic pass over the packing-list table
    vs the prints. Auto-applies confident fixes:
      • remakes — '242R' gets highlighted at 242's location
      • single-candidate OCR-misread renames, validated against the DXF panel
        list (the bad number must NOT exist in CAD; the corrected one must
        exist, be unused, and already be located on the prints)
    Ambiguous cases come back as suggestions; the rest with a reason. Every
    applied fix is recorded in corrections.json so re-scans keep it."""
    state = _pl_load_state(job_name)
    if not state:
        return jsonify({"error": "No delivery data yet — process a packing list first."}), 404
    locations = _pl_load_locations(job_name)

    # Panel Mapper locations widen the net (same merge the editor uses)
    pm_sess = _pm_load_session(job_name)
    pm_locs = {}
    if pm_sess:
        lp = pm_sess.get("locs", "")
        if lp and os.path.isfile(lp):
            try:
                with open(lp) as f:
                    pm_locs = _json.load(f)
            except Exception:
                pm_locs = {}

    def find_loc(panel):
        """Location entry for a panel: exact key, then duplicate/label keys."""
        for src in (locations, pm_locs):
            if panel in src and isinstance(src[panel], dict):
                return src[panel]
            for k, v in src.items():
                if not isinstance(v, dict):
                    continue
                if k.split("#")[0] == panel or v.get("label") == panel:
                    return v
        return None

    dxf_set = None
    dxf_dir = _pl_dxf_dir(job_name)
    if dxf_dir:
        try:
            from packing_list_engine import _load_dxf_panel_set
            dxf_set = _load_dxf_panel_set(dxf_dir)
        except Exception:
            dxf_set = None

    # Digit pairs OCR commonly confuses on KPS packing lists
    _CONF = {"0": "689", "1": "47", "2": "7", "3": "8", "4": "1", "5": "68",
             "6": "058", "7": "12", "8": "0356", "9": "0345"}

    def _confusable(a, b):
        """True if a→b is a plausible single-digit misread (same length)."""
        if len(a) != len(b):
            return False
        diff = [(x, y) for x, y in zip(a, b) if x != y]
        if len(diff) != 1:
            return False
        x, y = diff[0]
        return y in _CONF.get(x, "") or x in _CONF.get(y, "")

    placed, renamed, suggestions, unresolved, added = [], [], [], [], []
    corr = _pl_load_corrections(job_name)

    # Stage 0: panels PRINTED on a packing list that never made it into the
    # delivery table (parse misses — e.g. a short panel row the parser skipped).
    # Sweeps the editor's OCR position scans (pl_pos_v4_* caches; uncached
    # lists are scanned now, slower on first run). A number only counts if CAD
    # or the prints confirm it's a real panel — OCR junk is skipped. Panels the
    # user deleted before are never resurrected.
    pl_dir = safe_join(job_name, "Packing Lists")
    if os.path.isdir(pl_dir):
        try:
            from packing_list_engine import scan_packing_list_positions
            for fname in sorted(os.listdir(pl_dir)):
                if not fname.lower().endswith(".pdf"):
                    continue
                label = fname[:-4]
                safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", fname)
                cache = os.path.join(_pl_tracking_dir(job_name),
                                     f"pl_pos_v4_{safe_name}.json")
                os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
                try:
                    data = scan_packing_list_positions(
                        os.path.join(pl_dir, fname), cache_path=cache)
                except Exception:
                    continue
                for pos in (data or {}).get("positions", []):
                    p = str(pos.get("panel", "")).strip()
                    if not p or p in state or p in corr.get("deletions", []):
                        continue
                    loc = find_loc(p)
                    if not ((dxf_set and p in dxf_set) or loc):
                        continue            # not a real panel — OCR junk
                    state[p] = {"skid": "", "shipment": label}
                    if (loc and p not in locations and loc.get("bbox")
                            and loc.get("page") is not None):
                        locations[p] = {"page": int(loc["page"]), "bbox": loc["bbox"]}
                    corr["additions"][p] = {
                        "panel": p, "skid": "", "shipment": label,
                        "page": (int(loc["page"]) if loc and loc.get("page") is not None else None),
                        "bbox": (loc.get("bbox") if loc else None)}
                    _log_global_correction(job_name, {"type": "auto-fix-add",
                                                      "panel": p, "shipment": label})
                    added.append({"panel": p, "shipment": label, "located": bool(loc)})
        except Exception:
            pass

    missed = [p for p in state if not find_loc(p)]

    for p in missed:
        # (a) remake: '242R' sits where '242' is
        m = re.match(r"^(\d+)R$", p)
        if m:
            loc = find_loc(m.group(1))
            if loc and loc.get("bbox") and loc.get("page") is not None:
                locations[p] = {"page": int(loc["page"]), "bbox": loc["bbox"], "label": p}
                corr["additions"][p] = {"panel": p, "skid": state[p].get("skid", ""),
                                        "shipment": state[p].get("shipment", ""),
                                        "page": int(loc["page"]), "bbox": loc["bbox"]}
                _log_global_correction(job_name, {"type": "auto-fix-remake", "panel": p,
                                                  "page": int(loc["page"]), "bbox": loc["bbox"]})
                placed.append({"panel": p, "how": f"remake of {m.group(1)}"})
                continue

        # (b) OCR-misread rename — only when CAD confirms the number is wrong
        if dxf_set and p not in dxf_set and re.match(r"^\d+$", p):
            cands = [d for d in sorted(dxf_set)
                     if d not in state and _confusable(p, d) and find_loc(d)]
            if len(cands) == 1:
                new = cands[0]
                state[new] = state.pop(p)
                loc = find_loc(new)
                if new not in locations:
                    locations[new] = {"page": int(loc["page"]), "bbox": loc["bbox"]}
                corr["renames"][p] = new
                _log_global_correction(job_name, {"type": "auto-fix-rename",
                                                  "from": p, "to": new})
                renamed.append({"from": p, "to": new})
                continue
            if len(cands) > 1:
                suggestions.append({"panel": p, "candidates": cands[:6],
                                    "reason": "number not in CAD — several close matches"})
                continue
            unresolved.append({"panel": p,
                               "reason": "number not in CAD and no close match on the prints"})
            continue

        if dxf_set and p in dxf_set:
            unresolved.append({"panel": p,
                               "reason": "in CAD but its page isn't mapped — place by hand"})
        else:
            unresolved.append({"panel": p,
                               "reason": "no CAD data to locate it — place by hand"})

    changed = bool(placed or renamed or added)
    regen_ok, regen_err = changed, None
    if changed:
        _pl_save_corrections(job_name, corr)
        os.makedirs(_pl_tracking_dir(job_name), exist_ok=True)
        with open(_pl_state_path(job_name), "w") as f:
            _json.dump(state, f)
        with open(_pl_cache_path(job_name), "w") as f:
            _json.dump(locations, f)
        # Regenerate the tracked PDF — same path as manual Save in the editor
        try:
            from packing_list_engine import (generate_tracked_blueprint,
                                             generate_tracked_blueprint_panel_map)
            ship_colors = _pl_assign_colors(job_name, state)
            pm_locs_path = pm_sess.get("locs", "") if pm_sess else ""
            pm_scan_pdf = pm_sess.get("scan_pdf", "") if pm_sess else ""
            if (pm_sess and pm_locs_path and os.path.isfile(pm_locs_path)
                    and pm_scan_pdf and os.path.isfile(pm_scan_pdf)):
                with open(pm_locs_path) as f:
                    panel_locations_pm = _json.load(f)
                merged = dict(panel_locations_pm)
                merged.update(locations)
                generate_tracked_blueprint_panel_map(
                    pm_scan_pdf, merged, state, _pl_output_path(job_name),
                    shipment_colors=ship_colors)
            else:
                bp = _find_blueprint(job_name)
                if bp and locations:
                    generate_tracked_blueprint(
                        bp, state, locations, _pl_output_path(job_name),
                        shipment_colors=ship_colors)
        except Exception as e:
            regen_ok, regen_err = False, str(e)

    return jsonify({"ok": True, "added": added, "placed": placed, "renamed": renamed,
                    "suggestions": suggestions, "unresolved": unresolved,
                    "missed_before": len(missed),
                    "regenerated": regen_ok, "regen_error": regen_err})


@app.route("/api/pt/publish-prints/<path:job_name>", methods=["POST"])
@login_required
def pt_publish_prints(job_name):
    """Review-tab Publish: merge the marked-up (delivery-tracked) pages back into
    the FULL print set — each marked page replaces its original page. The result
    is saved under a stable name in Blueprints (overwritten each publish) and the
    public 'Marked-Up Prints' link is pointed at it (the LINK never changes)."""
    import shutil
    tracked = _pl_output_path(job_name)
    if not os.path.exists(tracked):
        return jsonify({"error": "No marked-up prints to publish yet"}), 404
    bp_dir = safe_join(job_name, "Blueprints")
    os.makedirs(bp_dir, exist_ok=True)
    dest = os.path.join(bp_dir, f"{job_name} - Marked-Up Prints.pdf")

    sess = _pm_load_session(job_name)
    src = (sess or {}).get("src_pdf") or ""
    pages = (sess or {}).get("pages") or []
    try:
        if sess and src and os.path.isfile(src) and pages:
            # tracked page i corresponds to sorted(pages)[i] (scan order)
            _pm_merge_full(src, tracked, pages, dest)
        else:
            # No Panel Mapper session: the tracked doc already covers the full set
            shutil.copy2(tracked, dest)
    except Exception as e:
        return jsonify({"error": f"Could not build the full print set: {e}"}), 500

    # Swap the document behind the public Marked-Up Prints link (same link;
    # created on first publish)
    data = _pub_load()
    info = data.setdefault(job_name, {}).setdefault("prints", {})
    info["file"] = f"Blueprints/{os.path.basename(dest)}"
    if not info.get("token"):
        info["token"] = secrets.token_urlsafe(16)
    now = datetime.now()
    info["published"] = f"{now.month}/{now.day}/{now.year % 100} {now:%H:%M}"
    _pub_save(data)
    return jsonify({"ok": True, "filename": os.path.basename(dest),
                    "url": request.url_root.rstrip("/") + "/pl/" + info["token"]})


@app.route("/pl/<token>/panels")
def public_link_panels(token):
    """Panel locations for the public Panels-Only viewer (?panel= zoom).
    Same merged map the Review editor uses — page indices line up with the
    Panel Mapper panels_only.pdf. Token-gated like every public route."""
    job, slot, info = _pub_find(token)
    if not job or slot != "panels_only":
        abort(404)
    locations = _pl_load_locations(job)
    pm_sess = _pm_load_session(job)
    if pm_sess:
        pm_locs_path = pm_sess.get("locs", "")
        if pm_locs_path and os.path.isfile(pm_locs_path):
            try:
                with open(pm_locs_path) as f:
                    pm_locs = json.load(f)
                for panel, loc in pm_locs.items():
                    if panel not in locations:
                        locations[panel] = loc
            except Exception:
                pass
    return jsonify(locations)


# ── Field Report → Discord ───────────────────────────────────
# Each channel maps to ONE Discord webhook URL (created per channel in
# Discord: Edit Channel → Integrations → Webhooks → New Webhook → Copy URL).
# URLs live in <JOBS_DIR>/.discord_webhooks.json — on the persistent jobs
# volume (like .public_links.json) so they survive deploys and stay out of git.
DISCORD_CHANNELS = [
    ("gc",           "GC"),
    ("pm",           "PM"),
    ("todo",         "To-Do"),
    ("safety",       "Safety"),
    ("progress",     "Progress"),
    ("extra_work",   "Extra Work"),
    ("trade_damage", "Trade Damage"),
]
_WEBHOOKS_FILE = os.path.join(JOBS_DIR, ".discord_webhooks.json")

def _load_webhooks():
    """Webhook config: {job: {channel_key: url}} — one Discord server per
    job. Legacy flat {channel: url} files load as Modesto Courthouse's
    channels (the original server belongs to that job)."""
    try:
        with open(_WEBHOOKS_FILE, encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return {}
    if raw and all(isinstance(v, str) for v in raw.values()):
        return {"Modesto Courthouse": raw}
    return {k: v for k, v in raw.items() if isinstance(v, dict)}

# Sent reports awaiting a "Mark task complete" click. Keyed by token; each
# record keeps the sent embed + Discord message id so the completion route can
# PATCH the message in place. Persisted on the jobs volume like the webhooks.
_FIELD_REPORTS_FILE = os.path.join(JOBS_DIR, ".field_reports.json")

def _fr_load_reports():
    try:
        with open(_FIELD_REPORTS_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _fr_save_reports(reports):
    # Cap growth: keep the newest 400 reports
    if len(reports) > 400:
        oldest = sorted(reports, key=lambda t: reports[t].get("created", ""))
        for t in oldest[:len(reports) - 400]:
            del reports[t]
    with open(_FIELD_REPORTS_FILE, "w", encoding="utf-8") as f:
        json.dump(reports, f, indent=1)

def _fr_to_jpeg(data):
    """Browsers can't decode iPhone HEIC photos, so the page sends those as
    originals — convert any non-JPEG upload to a resized JPEG here.
    Returns JPEG bytes, or the input unchanged if conversion isn't possible."""
    if data[:3] == b"\xff\xd8\xff":          # already JPEG (client-compressed)
        return data
    try:
        try:
            import pillow_heif
            pillow_heif.register_heif_opener()
        except ImportError:
            pass
        from PIL import Image as _Img
        im = _Img.open(io.BytesIO(data))
        im = im.convert("RGB")
        im.thumbnail((1600, 1600))
        out = io.BytesIO()
        im.save(out, "JPEG", quality=82)
        return out.getvalue()
    except Exception:
        return data

def _fr_display_name(email):
    """caseyc@pacificerectors.com -> 'Casey C.' (first name + last initial)."""
    prefix = (email or "").split("@")[0]
    if email.endswith("@pacificerectors.com") and len(prefix) > 2:
        return prefix[:-1].capitalize() + " " + prefix[-1].upper() + "."
    return prefix.capitalize() or "Unknown"

@app.route("/field-report")
@login_required
def field_report():
    return render_template("field_report.html", page_title="Push to Discord")

# ── Voice note transcription (OpenAI) ────────────────────────
# Key lives on the jobs volume like the webhooks. When present, the tool's
# Dictate button records real audio and transcribes it server-side; when
# absent, the page falls back to the free browser Web Speech engine.
_TRANSCRIBE_KEY_FILE = os.path.join(JOBS_DIR, ".openai_key.txt")
# whisper-1, NOT gpt-4o-mini-transcribe: the 4o models hallucinate full wrong
# sentences on quiet/noisy jobsite audio; whisper degrades gracefully instead.
_TRANSCRIBE_MODEL = "whisper-1"
_TRANSCRIBE_PROMPT = ("Field note from a metal panel siding installation crew. Jobsite terms: "
                      "Pacific Erectors, KPS, fab sheet, packing list, skid, panel, girt, clip, "
                      "soffit, parapet, flashing, RFI, GC, foreman, blueprint, conex, gang box, "
                      "fire locker, fire extinguisher, gas can, shelf, scissor lift, boom lift, "
                      "lull, swing stage, safety harness, lanyard, tie-off.")
# Last recording + result, kept for debugging bad transcriptions
_FR_LAST_AUDIO = os.path.join(JOBS_DIR, ".last_voice_note.bin")
_FR_LAST_META  = os.path.join(JOBS_DIR, ".last_voice_note.json")

def _fr_normalize_audio(data):
    """Phone recordings (esp. iPhone Safari) often arrive 25dB too quiet for
    the transcriber. Decode, boost peak to -3dBFS, return 16kHz mono WAV.
    Returns None on any failure — caller falls back to the raw audio."""
    try:
        import io, wave
        import av
        import numpy as np
        c = av.open(io.BytesIO(data))
        st = c.streams.audio[0]
        sr = st.codec_context.sample_rate or 48000
        chunks = [fr.to_ndarray() for fr in c.decode(st)]
        if not chunks:
            return None
        a = np.concatenate(chunks, axis=-1).astype(np.float32)
        if a.ndim > 1:
            a = a.mean(axis=0)
        if np.abs(a).max() > 2:       # int16-scaled samples
            a = a / 32768.0
        peak = np.abs(a).max()
        if peak < 1e-6:               # pure silence
            return None
        a = np.clip(a * ((10 ** (-3 / 20)) / peak), -1, 1)
        idx = (np.arange(int(len(a) * 16000 / sr)) * (sr / 16000)).astype(int)
        pcm = (a[idx] * 32767).astype(np.int16)
        buf = io.BytesIO()
        with wave.open(buf, "wb") as w:
            w.setnchannels(1); w.setsampwidth(2); w.setframerate(16000)
            w.writeframes(pcm.tobytes())
        return buf.getvalue()
    except Exception:
        return None

def _load_transcribe_key():
    try:
        with open(_TRANSCRIBE_KEY_FILE, encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        return ""

@app.route("/api/field-report/config")
@login_required
def field_report_config():
    return jsonify({"transcribe": bool(_load_transcribe_key())})

@app.route("/api/field-report/transcribe", methods=["POST"])
@login_required
def field_report_transcribe():
    try:
        import requests
    except ImportError:
        return jsonify({"error": "Server missing the 'requests' library — run the full deploy (deploy.bat) once."}), 500
    key = _load_transcribe_key()
    if not key:
        return jsonify({"error": "Transcription isn’t set up yet (no API key on the server)."}), 400
    f = request.files.get("audio")
    if not f:
        return jsonify({"error": "No audio received."}), 400
    data = f.read()
    if len(data) < 1000:
        return jsonify({"error": "Recording was empty — try again."}), 400
    if len(data) > 20 * 1024 * 1024:
        return jsonify({"error": "Recording too long — keep voice notes under ~5 minutes."}), 413
    wav = _fr_normalize_audio(data)
    send = (wav, "note.wav", "audio/wav") if wav else \
           (data, f.filename or "note.webm", f.mimetype or "audio/webm")
    try:
        resp = requests.post(
            "https://api.openai.com/v1/audio/transcriptions",
            headers={"Authorization": f"Bearer {key}"},
            data={"model": _TRANSCRIBE_MODEL, "prompt": _TRANSCRIBE_PROMPT,
                  "language": "en", "temperature": 0},
            files={"file": (send[1], send[0], send[2])},
            timeout=60,
        )
    except requests.RequestException:
        return jsonify({"error": "Could not reach the transcription service — try again."}), 502

    text = resp.json().get("text", "") if resp.status_code == 200 else ""
    try:  # keep the last recording + result for debugging; never let this fail a request
        with open(_FR_LAST_AUDIO, "wb") as fa:
            fa.write(data)
        with open(_FR_LAST_META, "w", encoding="utf-8") as fm:
            json.dump({"filename": f.filename, "mimetype": f.mimetype, "bytes": len(data),
                       "mic": request.form.get("mic", ""),
                       "client_level_db": request.form.get("level", ""),
                       "normalized": bool(wav),
                       "status": resp.status_code, "model": _TRANSCRIBE_MODEL, "text": text,
                       "user": session.get("user", ""),
                       "time": datetime.now(timezone.utc).isoformat()}, fm)
    except Exception:
        pass

    if resp.status_code == 200:
        return jsonify({"text": text})
    if resp.status_code == 401:
        return jsonify({"error": "Transcription API key is invalid — check the key on the server."}), 502
    return jsonify({"error": f"Transcription failed (HTTP {resp.status_code})."}), 502

@app.route("/api/field-report/last-audio")
@login_required
def field_report_last_audio():
    """Debug: download the most recent voice recording (?info=1 for metadata)."""
    if request.args.get("info"):
        try:
            with open(_FR_LAST_META, encoding="utf-8") as fm:
                return jsonify(json.load(fm))
        except Exception:
            return jsonify({"error": "No recording captured yet."}), 404
    if not os.path.isfile(_FR_LAST_AUDIO):
        return jsonify({"error": "No recording captured yet."}), 404
    return send_file(_FR_LAST_AUDIO, download_name="last_voice_note.bin")

@app.route("/api/field-report/channels")
@login_required
def field_report_channels():
    hooks = _load_webhooks()
    return jsonify({
        "channels": [{"key": k, "label": l} for k, l in DISCORD_CHANNELS],
        "jobs": {job: {k: bool(chans.get(k)) for k, _ in DISCORD_CHANNELS}
                 for job, chans in hooks.items()},
    })

@app.route("/api/field-report/send", methods=["POST"])
@login_required
def field_report_send():
    try:
        import requests
    except ImportError:
        return jsonify({"error": "Server missing the 'requests' library — run the full deploy (deploy.bat) once."}), 500

    hooks = _load_webhooks()
    labels = dict(DISCORD_CHANNELS)

    job     = (request.form.get("job") or "").strip()
    channel = (request.form.get("channel") or "").strip()
    note    = (request.form.get("note") or "").strip()
    photos  = request.files.getlist("photos")

    if job not in hooks:
        return jsonify({"error": "Pick a job first."}), 400
    if channel not in labels:
        return jsonify({"error": "Unknown channel."}), 400
    webhook = hooks[job].get(channel)
    if not webhook:
        return jsonify({"error": f"The “{labels[channel]}” channel isn’t set up for {job} yet (no webhook configured)."}), 400
    if not note and not photos:
        return jsonify({"error": "Add a photo or a note before sending."}), 400
    if len(photos) > 10:
        return jsonify({"error": "Discord allows up to 10 photos per message."}), 400

    poster = session.get("user", "")
    # "Checkbox": a link in the message that marks the task complete when
    # tapped (webhook messages can't carry real Discord buttons — that needs
    # a full bot app). Tapping it hits /fr/<token>, which edits this message.
    # ONLY To-Do messages are tasks — other channels post without it.
    token = checkbox_md = ""
    if channel == "todo":
        token = secrets.token_urlsafe(12)
        checkbox_md = f"[☐ **Mark task complete**]({request.url_root.rstrip('/')}/fr/{token})"
    # Urgent (To-Do only): red embed + siren title, both in Discord and on
    # the web board. Completing the task flips it green like any other.
    urgent = channel == "todo" and request.form.get("urgent") == "1"
    desc = "\n\n".join(p for p in (note[:3800], checkbox_md) if p)
    embed = {
        "title": ("🚨 URGENT Field Report" if urgent else "📋 Field Report")
                 + (f" — {job}" if job != "Company" else ""),
        "color": 0xF85149 if urgent else 0x60B4F0,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    if desc:
        embed["description"] = desc
    if poster:
        embed["author"] = {"name": poster}

    payload = {"username": "PEI Field Report", "embeds": [embed]}
    files = []
    for i, f in enumerate(photos):
        data = f.read()
        if not data:
            continue
        data = _fr_to_jpeg(data)
        files.append((f"files[{i}]", (f"photo{i + 1}.jpg", data, "image/jpeg")))

    try:
        resp = requests.post(
            webhook,
            params={"wait": "true"},  # make Discord return the message id
            data={"payload_json": json.dumps(payload)},
            files=files or None,
            timeout=30,
        )
    except requests.RequestException:
        return jsonify({"error": "Could not reach Discord. Check the connection and try again."}), 502

    if resp.status_code in (200, 204):
        try:
            if not token:           # no checkbox -> nothing to track
                return jsonify({"ok": True, "channel": labels[channel]})
            message_id = resp.json()["id"]
            reports = _fr_load_reports()
            reports[token] = {
                "job": job,
                "channel": channel,
                "message_id": message_id,
                "embed": embed,
                "checkbox_md": checkbox_md,
                "created": datetime.now(timezone.utc).isoformat(),
                "completed_by": None,
            }
            _fr_save_reports(reports)
        except Exception:
            pass  # report still sent; only the complete-link is dead
        return jsonify({"ok": True, "channel": labels[channel]})
    if resp.status_code == 413:
        return jsonify({"error": "Photos are too large for Discord. Try sending fewer at once."}), 413
    return jsonify({"error": f"Discord rejected the message (HTTP {resp.status_code})."}), 502



# No login on these routes — anyone with the (unguessable) link can complete
# the task, same trust model as Public Links. The GET page changes nothing;
# the page's own JS fires the POST, so link-preview crawlers (Discord, iOS,
# mail scanners) that fetch the URL can't accidentally complete tasks.
_FR_CLICK_PAGE = """<!DOCTYPE html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Task | PEI Tools</title><link rel="icon" href="/static/favicon.png">
<style>body{font-family:'Inter',-apple-system,sans-serif;background:#0d1117;color:#e6edf3;
display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;padding:20px}
.box{background:#1c2230;border:1px solid #30363d;border-radius:14px;padding:36px 32px;
max-width:420px;text-align:center}.big{font-size:2.6rem;margin-bottom:14px}
h1{font-size:1.15rem;margin:0 0 8px}p{color:#8b949e;font-size:.9rem;line-height:1.5;margin:0}</style>
</head><body><div class="box"><div class="big" id="i">⏳</div><h1 id="t">Marking task complete…</h1>
<p id="m"></p></div>
<script>
fetch('/fr/__TOKEN__/complete', {method: 'POST'})
  .then(r => r.json())
  .then(d => {
    if (d.already) { i.textContent = '✅'; t.textContent = 'Already completed';
      m.textContent = 'This task was already marked complete by ' + d.completed_by + '.'; }
    else if (d.ok) { i.textContent = '✅'; t.textContent = 'Task marked complete';
      m.textContent = 'The Discord message now shows it was completed by ' + d.by + '. You can close this tab.'; }
    else { i.textContent = '🤷'; t.textContent = 'Link not found';
      m.textContent = d.error || 'This task link is no longer on file.'; }
  })
  .catch(() => { i.textContent = '⚠️'; t.textContent = 'Something went wrong';
    m.textContent = 'Check your connection and tap the link again.'; });
</script></body></html>"""

@app.route("/fr/<token>")
def field_report_complete(token):
    return _FR_CLICK_PAGE.replace("__TOKEN__", token)

@app.route("/fr/<token>/complete", methods=["POST"])
def field_report_complete_post(token):
    reports = _fr_load_reports()
    rec = reports.get(token)
    if not rec:
        return jsonify({"error": "This task link is no longer on file."}), 404
    if rec.get("completed_by"):
        return jsonify({"already": True, "completed_by": rec["completed_by"]})

    user = session.get("user", "")
    name = _fr_display_name(user) if user else "a crew member"
    done_md = (f"✅ **Completed by {name}** · "
               f"<t:{int(datetime.now(timezone.utc).timestamp())}:f>")
    embed = rec["embed"]
    embed["description"] = embed["description"].replace(rec["checkbox_md"], done_md)
    embed["color"] = 0x3FB950  # green once complete
    if "title" in embed:       # an urgent task stops being urgent when done
        embed["title"] = embed["title"].replace("🚨 URGENT Field Report", "📋 Field Report")

    hooks = _load_webhooks()
    rec_job = rec.get("job", "")
    if rec_job not in hooks and len(hooks) == 1:
        rec_job = next(iter(hooks))  # heal records from before the server was renamed
    webhook = hooks.get(rec_job, {}).get(rec["channel"], "")
    if webhook:
        try:
            import requests
            requests.patch(f"{webhook}/messages/{rec['message_id']}",
                           json={"embeds": [embed]}, timeout=15)
        except Exception:
            pass  # completion still recorded on our side

    rec["completed_by"] = name
    rec["completed_at"] = datetime.now(timezone.utc).isoformat()
    _fr_save_reports(reports)
    return jsonify({"ok": True, "by": name})


# ── To-Do channel web viewer ─────────────────────────────────
# Public, token-gated page (Public Links trust model) showing the #to-do
# Discord channel. Reading a channel needs a BOT token (webhooks can't read);
# it lives on the jobs volume like the other secrets. The channel id is
# resolved once from the to-do webhook itself.
_DISCORD_BOT_FILE = os.path.join(JOBS_DIR, ".discord_bot_token.txt")
_TODO_VIEW_TOKENS_FILE = os.path.join(JOBS_DIR, ".todo_view_tokens.json")
_todo_cache = {}  # job -> {"cid", "msgs", "t"} (one board per job server)

def _load_bot_token():
    try:
        with open(_DISCORD_BOT_FILE, encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        return ""

def _todo_tokens():
    try:
        with open(_TODO_VIEW_TOKENS_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _todo_token_for(job, create=False):
    toks = _todo_tokens()
    if toks.get(job):
        return toks[job]
    if not create:
        return ""
    toks[job] = secrets.token_urlsafe(16)
    with open(_TODO_VIEW_TOKENS_FILE, "w", encoding="utf-8") as f:
        json.dump(toks, f, indent=2)
    return toks[job]

def _todo_job_for(tok):
    if not tok:
        return None
    for job, t in _todo_tokens().items():
        if t == tok:
            return job
    return None

def _todo_channel_id(job):
    ent = _todo_cache.setdefault(job, {})
    if ent.get("cid"):
        return ent["cid"]
    import requests
    webhook = _load_webhooks().get(job, {}).get("todo", "")
    if not webhook:
        return None
    try:
        cid = requests.get(webhook, timeout=15).json().get("channel_id")
    except Exception:
        return None
    ent["cid"] = cid
    return cid

@app.route("/todo")
@login_required
def todo_link():
    """Logged-in helper: lists every job's to-do board with its share link."""
    hooks = _load_webhooks()
    rows = []
    for job in sorted(hooks, key=lambda j: (j != "Company", j)):
        if hooks[job].get("todo"):
            try:
                tok = _todo_token_for(job, create=True)
            except Exception:
                continue
            rows.append(f'<li style="margin:8px 0"><a style="color:#60b4f0" '
                        f'href="/todo/{tok}">{job}</a></li>')
    body = ("<ul style='list-style:none'>" + "".join(rows) + "</ul>") if rows else \
           "<p>No to-do channels configured yet.</p>"
    return ('<!DOCTYPE html><html><head><meta charset="utf-8"><title>To-Do Boards | PEI Tools</title>'
            '<meta name="viewport" content="width=device-width, initial-scale=1.0">'
            '<style>body{font-family:Inter,sans-serif;background:#0d1117;color:#e6edf3;padding:40px 20px;'
            'max-width:480px;margin:0 auto}h1{font-size:1.2rem}p,li{font-size:.95rem}</style></head>'
            f'<body><h1>To-Do Boards</h1>{body}</body></html>')

@app.route("/todo/<tok>")
def todo_view(tok):
    job = _todo_job_for(tok)
    if not job:
        abort(404)
    return render_template("todo_view.html", tok=tok, job=job)

@app.route("/api/todo/<tok>/messages")
def todo_messages(tok):
    job = _todo_job_for(tok)
    if not job:
        abort(404)
    import time as _t
    ent = _todo_cache.setdefault(job, {})
    if ent.get("msgs") is not None and _t.time() - ent.get("t", 0) < 20:
        return jsonify(ent["msgs"])  # cache: don't hammer Discord on refreshes
    bot = _load_bot_token()
    if not bot:
        return jsonify({"error": "The viewer isn’t set up yet (no Discord bot token on the server)."}), 503
    cid = _todo_channel_id(job)
    if not cid:
        return jsonify({"error": f"No to-do channel is set up for {job} yet."}), 503
    import requests
    try:
        r = requests.get(f"https://discord.com/api/v10/channels/{cid}/messages",
                         params={"limit": 50},
                         headers={"Authorization": f"Bot {bot}"}, timeout=20)
    except requests.RequestException:
        return jsonify({"error": "Could not reach Discord — try again."}), 502
    if r.status_code == 401:
        return jsonify({"error": "The bot token is invalid — check it on the server."}), 503
    if r.status_code == 403:
        return jsonify({"error": "The bot doesn’t have access to #to-do — invite it to the server with View Channels + Read Message History."}), 503
    if r.status_code != 200:
        return jsonify({"error": f"Discord error (HTTP {r.status_code})."}), 502
    msgs = []
    for m in r.json():  # Discord returns newest first
        msgs.append({
            "author": m.get("author", {}).get("username", "?"),
            "time": m.get("timestamp", ""),
            "content": m.get("content", ""),
            "embeds": [{"title": e.get("title", ""),
                        "description": e.get("description", ""),
                        "color": e.get("color", 0)} for e in m.get("embeds", [])],
            "images": [a["url"] for a in m.get("attachments", [])
                       if (a.get("content_type") or "").startswith("image/")],
            "files": [{"name": a.get("filename", "file"), "url": a["url"]}
                      for a in m.get("attachments", [])
                      if not (a.get("content_type") or "").startswith("image/")],
        })
    ent["msgs"] = msgs
    ent["t"] = _t.time()
    return jsonify(msgs)


if __name__ == "__main__":
    app.run(debug=True)
